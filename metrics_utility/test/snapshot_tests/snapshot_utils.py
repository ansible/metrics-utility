from pprint import pprint
from dataclasses import dataclass
from typing import Dict, List
import json
from pathlib import Path
import os
import sys
import subprocess
import copy
import re
import shutil
import openpyxl
import openpyxl.utils

import warnings
warnings.filterwarnings("ignore", category=ResourceWarning)

@dataclass
class DataShape:
    env_vars: Dict[str, str]
    params: List[str]
    # values:
    # run_command: 'Yes' - runs command using subprocess, 
    # in future we may want to run function directly due to mocking datetime

    # generated: date when report was generated
    custom_params: Dict[str, str]

def create_directory_if_not_exists(directory_path):
    try:
        os.makedirs(directory_path, exist_ok=True)
    except Exception as e:
        print(f"Error creating directory: {e}")

def parse_json_file(file_path):
    try:
        with open(file_path, 'r') as file:
            data = json.load(file) 
            return data
    except json.JSONDecodeError as e:
        print(f"Error decoding JSON from file {file_path}: {e}")
    except FileNotFoundError:
        print(f"File not found: {file_path}")
    except Exception as e:
        print(f"An error occurred: {e}")
    return None

def save_snapshot_definition(data: DataShape, path: str):
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        # Open the file in write mode
        with open(path, 'w') as json_file:
            # Serialize the dictionary into JSON and write to the file
            json.dump(data, json_file, indent=4)
            json_file.flush()  # Explicitly flush the buffer
        print(f"Data successfully saved to {path}")
    except Exception as e:
        print(f"An error occurred while saving the data: {e} into {path}")

def find_json_files(directory):
    return list(Path(directory).rglob("*.json"))

def run_and_generate_snapshot_definitions(directory):
    print(f'\nRun and generate snapshots from {directory}\n')
    json_files = find_json_files(directory)

    for json_file in json_files:
        print(f'Found {json_file}')
        output_dir = json_file.as_posix().removesuffix('.json')
        create_directory_if_not_exists(output_dir)
        
        data = parse_json_file(json_file)
        generated_file = run_snapshot_definition(data)
        output_file = output_dir + "/report.xlsx"

        if os.path.exists(output_file):
            os.remove(output_file)

        shutil.move(generated_file, output_file)
        print(f'Report generated and moved from {generated_file} to {output_file}\n')
    

# Returns path to generated file
def run_snapshot_definition(data):
    env_vars = copy.deepcopy(data['env_vars'])
    env_vars["AWX_LOGGING_MODE"] = 'stdout'
    python_executable = sys.executable
    
    params = []
    params.append(python_executable)
    params.extend(data['params'])

    # Runs command, in future, we want to support also calling the test function directly due to mocking datetime.now
    if (data['custom_params']['run_command'] == 'Yes'):
        result = subprocess.run(
            params,
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            env=env_vars,
        )

        if result.returncode != 0:
            print('Generating of report failed')
            print(result.stderr)
            return
        
        text = result.stderr + '/n' + result.stdout
    else:
        return ''
    
    generated_file = get_file_name(params, env_vars)

    if generated_file:
        return generated_file
        
    # Regular expression to capture the file path if get_file_name was not able to compute it
    pattern = r"Report generated into directory:\s*([\w\-/]+)\.xlsx"

    match = re.search(pattern, text)
    if match:
        return match.group(1) + '.xlsx'

    return ''

def run_and_test_snapshot_definitions(directory):
    json_files = find_json_files(directory)
    print('Found definitions:')
    pprint(json_files)
    for json_file in json_files:
        print('')
        data_dir = json_file.as_posix().removesuffix('.json')
        data : DataShape = parse_json_file(json_file)

        original_file = './' + data_dir + "/report.xlsx"
        generated_file = run_snapshot_definition(data)

        print(f'Compare {original_file} to {generated_file}')
        # compare the generated and original_file
            
        if (data["env_vars"]["METRICS_UTILITY_REPORT_TYPE"] == 'CCSPv2'):
            compare_ccspv2_reports(original_file, generated_file)

        if (data["env_vars"]["METRICS_UTILITY_REPORT_TYPE"] == 'CCSP'):
            compare_ccsp_reports(original_file, generated_file)
    

        if os.path.exists(generated_file):
            print(f'Removing {generated_file}')

def compare_ccspv2_reports(original_report_path, generated_report_path):
    print(f'Opening {generated_report_path}')
    g_wb = openpyxl.load_workbook(filename = generated_report_path)  
                      
    print(f'Opening {original_report_path}')
    o_wb = openpyxl.load_workbook(filename = original_report_path)

    try:
        compare_worksheets(g_wb, o_wb, 0, ['H1', 'B5'])
        compare_worksheets(g_wb, o_wb, 1, [])
        compare_worksheets(g_wb, o_wb, 2, [])
    finally:
        g_wb.close()
        o_wb.close()

def compare_ccsp_reports(original_report_path, generated_report_path):
    print(f'Opening {generated_report_path}')
    g_wb = openpyxl.load_workbook(filename = generated_report_path)  
                      
    print(f'Opening {original_report_path}')
    o_wb = openpyxl.load_workbook(filename = original_report_path)

    try:
        compare_worksheets(g_wb, o_wb, 0, ['B5'])
        compare_worksheets(g_wb, o_wb, 1, [])
    finally:
        g_wb.close()
        o_wb.close()

def compare_worksheets(workbook_generated, workbook_original, sheet_number, exceptions : List[str]):   
    worksheet_generated = workbook_generated.worksheets[sheet_number]
    worksheet_original = workbook_original.worksheets[sheet_number]
    
    max_row_1 = worksheet_original.max_row
    max_column_1 = worksheet_original.max_column

    max_row_2 = worksheet_generated.max_row
    max_column_2 = worksheet_generated.max_column

    assert (
            max_column_1 == max_column_2
    ), f"Number of columns do not match for sheet number: {sheet_number}. Actual value = {max_column_2}, expected value = {max_column_1}"

    assert (
            max_row_1 == max_row_2
    ), f"Number of rows do not match for sheet number: {sheet_number}. Actual value = {max_row_2}, expected value = {max_row_1}"


    for column in range(1, max_column_1+1):
        for row in range(1, max_row_1+1):
            addr = openpyxl.utils.get_column_letter(column) + str(row)
            
            if addr not in exceptions:
                val_g = worksheet_generated[addr].value
                val_o = worksheet_original[addr].value 
               
                assert (
                    val_g == val_o
                    ), f"Column names do not match for sheet number: {sheet_number}. Address {addr}. Actual value = {val_g}, expected value = {val_o}"
    
def get_file_name(params, env_vars):
    if env_vars['METRICS_UTILITY_REPORT_TYPE'] == 'CCSPv2':
        month = get_param_value(params, 'month')
        if month:
            return f'CCSPv2-{month}.xlsx'
        
        # since and until has weirdly generated names, use the file name from terminal output of command
    return None

def get_param_value(params : List[str], name):
    for param in params:
        if param.startswith(name + '='):
            key, value = param.split('=')
            if key == name:
                return value
    return None

def get_entry_point_directory():
    # Get the absolute path of the entry-point file
    entry_point_file = os.path.abspath(sys.argv[0])
    # Get the directory containing the entry-point file
    entry_point_dir = os.path.dirname(entry_point_file)
    return entry_point_dir