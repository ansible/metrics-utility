"""Unit tests for automation_controller_billing/report/base.py helper methods.

All helpers under test are pure or depend only on extra_params / a small
DataFrame – no database or file I/O required.
"""

import pandas as pd

from metrics_utility.automation_controller_billing.report.base import Base
from metrics_utility.metric_utils import INDIRECT as _INDIRECT


# ---------------------------------------------------------------------------
# Minimal concrete subclass
# ---------------------------------------------------------------------------


class ConcreteReport(Base):
    """Instantiable subclass for testing Base helpers."""

    def __init__(self, extra_params=None):
        from openpyxl import Workbook

        self.wb = Workbook()
        self.extra_params = extra_params or {}
        self.dataframes = {}


def _report(dedup=False):
    deduplicator = 'ccsp-experimental' if dedup else None
    return ConcreteReport(extra_params={'deduplicator': deduplicator, 'optional_sheets': None})


# ---------------------------------------------------------------------------
# has_dedup_enabled
# ---------------------------------------------------------------------------


class TestHasDedupEnabled:
    def test_returns_true_when_ccsp_experimental(self):
        assert _report(dedup=True).has_dedup_enabled() is True

    def test_returns_false_when_no_deduplicator(self):
        assert _report(dedup=False).has_dedup_enabled() is False

    def test_returns_false_for_other_deduplicator_name(self):
        r = ConcreteReport(extra_params={'deduplicator': 'ccsp', 'optional_sheets': None})
        assert r.has_dedup_enabled() is False


# ---------------------------------------------------------------------------
# convert_cell
# ---------------------------------------------------------------------------


class TestConvertCell:
    def test_scalar_string_returned_unchanged(self):
        assert _report().convert_cell('hello') == 'hello'

    def test_scalar_int_returned_unchanged(self):
        assert _report().convert_cell(42) == 42

    def test_none_returned_unchanged(self):
        assert _report().convert_cell(None) is None

    def test_set_converted_to_sorted_json_array(self):
        import json

        result = _report().convert_cell({'b', 'a', 'c'})
        parsed = json.loads(result)
        assert parsed == ['a', 'b', 'c']

    def test_list_of_strings_sorted_and_json_encoded(self):
        import json

        result = _report().convert_cell(['c', 'a', 'b'])
        parsed = json.loads(result)
        assert parsed == ['a', 'b', 'c']

    def test_list_with_set_items_converted(self):
        import json

        result = _report().convert_cell([{'x', 'y'}])
        parsed = json.loads(result)
        assert isinstance(parsed[0], list)
        assert set(parsed[0]) == {'x', 'y'}

    def test_dict_with_set_values_serialised(self):
        import json

        result = _report().convert_cell({'key': {'b', 'a'}})
        parsed = json.loads(result)
        assert parsed['key'] == ['a', 'b']

    def test_dict_with_scalar_values_serialised(self):
        import json

        result = _report().convert_cell({'a': 1, 'b': 2})
        parsed = json.loads(result)
        assert parsed == {'a': 1, 'b': 2}

    def test_empty_set_produces_empty_json_array(self):
        import json

        result = _report().convert_cell(set())
        assert json.loads(result) == []

    def test_empty_dict_produces_empty_json_object(self):
        import json

        result = _report().convert_cell({})
        assert json.loads(result) == {}


# ---------------------------------------------------------------------------
# calculate_dedup_count
# ---------------------------------------------------------------------------


class TestCalculateDedupCount:
    def test_counts_elements_in_set(self):
        r = _report()
        series = pd.Series([{1, 2, 3}, {4}])
        result = r.calculate_dedup_count(series)
        assert list(result) == [3, 1]

    def test_counts_elements_in_list(self):
        r = _report()
        series = pd.Series([['a', 'b'], ['c']])
        result = r.calculate_dedup_count(series)
        assert list(result) == [2, 1]

    def test_scalar_counts_as_1(self):
        r = _report()
        series = pd.Series(['hostname'])
        result = r.calculate_dedup_count(series)
        assert list(result) == [1]

    def test_empty_set_counts_as_0(self):
        r = _report()
        series = pd.Series([set()])
        result = r.calculate_dedup_count(series)
        assert list(result) == [0]


# ---------------------------------------------------------------------------
# add_dedup_count_column
# ---------------------------------------------------------------------------


class TestAddDedupCountColumn:
    def test_adds_count_column_when_base_exists(self):
        r = _report()
        df = pd.DataFrame({'host_names_before_dedup': [{'h1', 'h2'}, {'h3'}]})
        result = r.add_dedup_count_column(df, 'host_names_before_dedup', 'host_names_before_dedup_count')
        assert 'host_names_before_dedup_count' in result.columns
        assert list(result['host_names_before_dedup_count']) == [2, 1]

    def test_no_op_when_base_column_absent(self):
        r = _report()
        df = pd.DataFrame({'other_col': [1, 2]})
        result = r.add_dedup_count_column(df, 'host_names_before_dedup', 'host_names_before_dedup_count')
        assert 'host_names_before_dedup_count' not in result.columns


# ---------------------------------------------------------------------------
# handle_dedup_columns_for_scope
# ---------------------------------------------------------------------------


class TestHandleDedupColumnsForScope:
    def test_adds_dedup_columns_when_enabled_and_column_present(self):
        r = _report(dedup=True)
        df = pd.DataFrame({'host_names_before_dedup': [{'h1'}, {'h2'}]})
        columns = ['host_name']
        convert_cols = []
        new_columns, new_convert = r.handle_dedup_columns_for_scope(df, columns, convert_cols)
        assert 'host_names_before_dedup' in new_columns
        assert 'host_names_before_dedup_count' in new_columns
        assert 'host_names_before_dedup' in new_convert

    def test_no_change_when_dedup_disabled(self):
        r = _report(dedup=False)
        df = pd.DataFrame({'host_names_before_dedup': [{'h1'}]})
        columns = ['host_name']
        convert_cols = []
        new_columns, new_convert = r.handle_dedup_columns_for_scope(df, columns, convert_cols)
        assert new_columns == ['host_name']
        assert new_convert == []

    def test_no_change_when_column_absent_even_if_dedup_enabled(self):
        r = _report(dedup=True)
        df = pd.DataFrame({'other': [1]})
        columns = ['host_name']
        convert_cols = []
        new_columns, _ = r.handle_dedup_columns_for_scope(df, columns, convert_cols)
        assert 'host_names_before_dedup' not in new_columns


# ---------------------------------------------------------------------------
# add_dedup_labels_if_needed
# ---------------------------------------------------------------------------


class TestAddDedupLabelsIfNeeded:
    def test_adds_labels_for_present_columns_when_dedup_enabled(self):
        r = _report(dedup=True)
        labels = {}
        result = r.add_dedup_labels_if_needed(labels, ['host_names_before_dedup'])
        assert 'host_names_before_dedup' in result

    def test_no_labels_added_when_dedup_disabled(self):
        r = _report(dedup=False)
        labels = {}
        result = r.add_dedup_labels_if_needed(labels, ['host_names_before_dedup'])
        assert result == {}

    def test_only_adds_labels_for_matching_columns(self):
        r = _report(dedup=True)
        labels = {}
        result = r.add_dedup_labels_if_needed(labels, ['host_names_before_dedup_count'])
        assert 'host_names_before_dedup_count' in result
        assert 'host_names_before_dedup' not in result


# ---------------------------------------------------------------------------
# convert_cell – additional branch (empty list)
# ---------------------------------------------------------------------------


class TestConvertCellEmptyList:
    def test_empty_list_produces_empty_json_array(self):
        import json

        result = _report().convert_cell([])
        assert json.loads(result) == []


# ---------------------------------------------------------------------------
# optional_report_sheets
# ---------------------------------------------------------------------------


class TestOptionalReportSheets:
    def test_returns_value_when_set(self):
        r = ConcreteReport(extra_params={'deduplicator': None, 'optional_sheets': ['SheetA', 'SheetB']})
        assert r.optional_report_sheets() == ['SheetA', 'SheetB']

    def test_returns_none_when_not_set(self):
        assert _report().optional_report_sheets() is None


# ---------------------------------------------------------------------------
# handle_dedup_aggregation
# ---------------------------------------------------------------------------


class TestHandleDedupAggregation:
    def test_adds_host_names_key_when_dedup_enabled(self):
        r = _report(dedup=True)
        agg = {}
        result = r.handle_dedup_aggregation(agg)
        assert 'host_names_before_dedup' in result
        assert result is agg  # mutates and returns the same dict

    def test_no_change_when_dedup_disabled(self):
        r = _report(dedup=False)
        agg = {'existing_key': ('col', 'sum')}
        result = r.handle_dedup_aggregation(agg)
        assert list(result.keys()) == ['existing_key']


# ---------------------------------------------------------------------------
# handle_dedup_columns_for_usage
# ---------------------------------------------------------------------------


class TestHandleDedupColumnsForUsage:
    def test_adds_columns_and_convert_when_enabled_and_col_present(self):
        r = _report(dedup=True)
        df = pd.DataFrame({'host_names_before_dedup': [{'h1', 'h2'}]})
        new_cols, new_convert = r.handle_dedup_columns_for_usage(df, ['host_name'], [])
        assert 'host_names_before_dedup' in new_cols
        assert 'host_names_before_dedup_count' in new_cols
        assert 'host_names_before_dedup' in new_convert

    def test_adds_columns_but_not_convert_when_enabled_and_col_absent(self):
        r = _report(dedup=True)
        df = pd.DataFrame({'other': [1]})
        new_cols, new_convert = r.handle_dedup_columns_for_usage(df, ['host_name'], [])
        assert 'host_names_before_dedup' in new_cols
        assert 'host_names_before_dedup' not in new_convert

    def test_no_change_when_dedup_disabled(self):
        r = _report(dedup=False)
        df = pd.DataFrame({'host_names_before_dedup': [{'h1'}]})
        new_cols, new_convert = r.handle_dedup_columns_for_usage(df, ['host_name'], [])
        assert new_cols == ['host_name']
        assert new_convert == []


# ---------------------------------------------------------------------------
# add_sheet / set_widths
# ---------------------------------------------------------------------------


class TestAddSheet:
    def test_creates_sheet_and_returns_it(self):
        r = _report()
        ws = r.add_sheet('MySheet', 1)
        assert ws.title == 'MySheet'
        assert r.wb.worksheets[1] is ws

    def test_sets_column_widths_when_provided(self):
        from openpyxl.utils import get_column_letter

        r = _report()
        ws = r.add_sheet('Wide', 1, widths={1: 20, 2: 35})
        assert ws.column_dimensions[get_column_letter(1)].width == 20
        assert ws.column_dimensions[get_column_letter(2)].width == 35

    def test_no_widths_argument_does_not_raise(self):
        r = _report()
        ws = r.add_sheet('Plain', 1)
        assert ws is not None


class TestSetWidths:
    def test_sets_multiple_column_widths(self):
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        r = _report()
        ws = Workbook().active
        r.set_widths(ws, {1: 15, 3: 25})
        assert ws.column_dimensions[get_column_letter(1)].width == 15
        assert ws.column_dimensions[get_column_letter(3)].width == 25


# ---------------------------------------------------------------------------
# _fix_event_host_names
# ---------------------------------------------------------------------------


def _mapping_df(original, install_uuid, job_remote_id, mapped_to):
    return pd.DataFrame(
        {
            'original_host_name': [original],
            'install_uuid': [install_uuid],
            'job_remote_id': [job_remote_id],
            'host_name': [mapped_to],
        }
    )


def _dest_df(host_name, install_uuid, job_remote_id):
    return pd.DataFrame(
        {
            'host_name': [host_name],
            'install_uuid': [install_uuid],
            'job_remote_id': [job_remote_id],
        }
    )


class TestFixEventHostNames:
    def test_returns_none_when_destination_is_none(self):
        r = _report()
        mapping = _mapping_df('old', 'u1', 'j1', 'new')
        assert r._fix_event_host_names(mapping, None) is None

    def test_maps_matching_host_name(self):
        r = _report()
        mapping = _mapping_df('old_host', 'u1', 'j1', 'new_host')
        dest = _dest_df('old_host', 'u1', 'j1')
        result = r._fix_event_host_names(mapping, dest)
        assert result['host_name'].iloc[0] == 'new_host'

    def test_adds_host_composite_id_column(self):
        r = _report()
        mapping = _mapping_df('old_host', 'u1', 'j1', 'new_host')
        dest = _dest_df('old_host', 'u1', 'j1')
        result = r._fix_event_host_names(mapping, dest)
        assert 'host_composite_id' in result.columns

    def test_unmapped_host_keeps_original_name(self):
        r = _report()
        mapping = _mapping_df('other_host', 'u1', 'j1', 'mapped')
        dest = _dest_df('unknown_host', 'u1', 'j1')
        result = r._fix_event_host_names(mapping, dest)
        assert result['host_name'].iloc[0] == 'unknown_host'


# ---------------------------------------------------------------------------
# Shared worksheet factory for _build_data_section_* tests
# ---------------------------------------------------------------------------


def _make_ws(r, title='TestSheet'):
    """Create and return a fresh worksheet in r's workbook."""
    r.wb.create_sheet(title=title)
    return r.wb.worksheets[-1]


# ---------------------------------------------------------------------------
# _build_data_section_scope
# ---------------------------------------------------------------------------


def _scope_df(**extra):
    data = {
        'host_name': ['host1', 'host2'],
        'last_automation': ['2024-01-01', '2024-01-02'],
        'organizations': [{'org1'}, {'org2'}],
        'inventories': [{'inv1'}, {'inv2'}],
        'canonical_facts': [{'cf': 'v1'}, {}],
        'facts': [{'f': 'v1'}, {}],
    }
    data.update(extra)
    return pd.DataFrame(data)


class TestBuildDataSectionScope:
    def test_returns_advanced_row_number(self):
        r = _report()
        ws = _make_ws(r)
        end_row = r._build_data_section_scope(1, ws, _scope_df())
        assert end_row > 1

    def test_header_row_is_bold(self):
        r = _report()
        ws = _make_ws(r)
        r._build_data_section_scope(1, ws, _scope_df())
        assert ws.cell(row=1, column=1).font.bold is True

    def test_data_rows_are_not_bold(self):
        r = _report()
        ws = _make_ws(r)
        r._build_data_section_scope(1, ws, _scope_df())
        assert ws.cell(row=2, column=1).font.bold is False

    def test_install_uuid_column_not_in_headers(self):
        r = _report()
        ws = _make_ws(r)
        r._build_data_section_scope(1, ws, _scope_df(install_uuid=['uid1', 'uid2']))
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        assert 'install_uuid' not in headers

    def test_dedup_column_label_present_when_enabled(self):
        r = _report(dedup=True)
        ws = _make_ws(r)
        df = _scope_df(host_names_before_dedup=[{'h1', 'h2'}, {'h3'}])
        r._build_data_section_scope(1, ws, df)
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        assert any(h and 'deduplication' in h for h in headers)

    def test_row_advance_equals_header_plus_data_rows(self):
        r = _report()
        ws = _make_ws(r)
        df = _scope_df()
        end_row = r._build_data_section_scope(1, ws, df)
        # dataframe_to_rows yields 1 header + len(df) data rows = 3 total
        assert end_row == 1 + 3


# ---------------------------------------------------------------------------
# _build_data_section_infrastructure_summary
# ---------------------------------------------------------------------------


def _infra_df_with_indirect():
    return pd.DataFrame(
        {
            'managed_node_type': [_INDIRECT, _INDIRECT, 0],
            'host_name': ['h1', 'h2', 'h3'],
            'facts': [
                {'infra_type': 'Cloud', 'infra_bucket': 'AWS', 'device_type': 'VM'},
                {'infra_type': 'Cloud', 'infra_bucket': 'AWS', 'device_type': 'Container'},
                {'infra_type': 'OnPrem', 'infra_bucket': 'Physical', 'device_type': 'Server'},
            ],
        }
    )


def _infra_df_no_indirect():
    return pd.DataFrame(
        {
            'managed_node_type': [0, 0],
            'host_name': ['h1', 'h2'],
            'facts': [{}, {}],
        }
    )


class TestBuildDataSectionInfrastructureSummary:
    def test_no_indirect_nodes_writes_message_and_advances_one_row(self):
        r = _report()
        ws = _make_ws(r)
        end_row = r._build_data_section_infrastructure_summary(1, ws, _infra_df_no_indirect())
        assert end_row == 2
        assert ws.cell(row=1, column=1).value == 'No indirect nodes found'

    def test_indirect_nodes_returns_row_beyond_start(self):
        r = _report()
        ws = _make_ws(r)
        end_row = r._build_data_section_infrastructure_summary(1, ws, _infra_df_with_indirect())
        assert end_row > 1

    def test_indirect_nodes_writes_column_headers(self):
        r = _report()
        ws = _make_ws(r)
        r._build_data_section_infrastructure_summary(1, ws, _infra_df_with_indirect())
        headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
        assert 'Infrastructure' in headers
        assert 'Device Type' in headers

    def test_facts_as_json_string_are_parsed(self):
        import json

        r = _report()
        ws = _make_ws(r)
        df = pd.DataFrame(
            {
                'managed_node_type': [_INDIRECT],
                'host_name': ['h1'],
                'facts': [json.dumps({'infra_type': 'Cloud', 'infra_bucket': 'GCP', 'device_type': 'VM'})],
            }
        )
        end_row = r._build_data_section_infrastructure_summary(1, ws, df)
        assert end_row > 1


# ---------------------------------------------------------------------------
# _build_data_section_usage_by_node
# ---------------------------------------------------------------------------


def _usage_by_node_df(**extra):
    data = {
        'host_name': ['host1', 'host1', 'host2'],
        'organization_name': ['org1', 'org1', 'org2'],
        'task_runs': [5, 3, 7],
        'first_automation': ['2024-01-01', '2024-01-02', '2024-01-03'],
        'last_automation': ['2024-01-02', '2024-01-03', '2024-01-04'],
        'managed_node_types_set': [['direct'], ['direct'], ['indirect']],
        'events': [['ev1'], ['ev2'], None],
        'canonical_facts': [{}, {}, {}],
        'facts': [{}, {}, {}],
    }
    data.update(extra)
    return pd.DataFrame(data)


class TestBuildDataSectionUsageByNode:
    def test_returns_advanced_row_number(self):
        r = _report()
        ws = _make_ws(r)
        end_row = r._build_data_section_usage_by_node(1, ws, _usage_by_node_df())
        assert end_row > 1

    def test_header_is_bold(self):
        r = _report()
        ws = _make_ws(r)
        r._build_data_section_usage_by_node(1, ws, _usage_by_node_df())
        assert ws.cell(row=1, column=1).font.bold is True

    def test_mode_by_organization_excludes_organizations_column(self):
        r = _report()
        ws = _make_ws(r)
        r._build_data_section_usage_by_node(1, ws, _usage_by_node_df(), mode='by_organization')
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        assert 'Automated by\norganizations' not in headers

    def test_managed_node_type_indirect_adds_extra_columns(self):
        r = _report()
        ws = _make_ws(r)
        r._build_data_section_usage_by_node(1, ws, _usage_by_node_df(), managed_node_type='indirect')
        headers = [ws.cell(row=1, column=c).value for c in range(1, 12)]
        assert 'Events' in headers

    def test_managed_node_type_direct_with_dedup_adds_facts_columns(self):
        r = _report(dedup=True)
        ws = _make_ws(r)
        df = _usage_by_node_df(host_names_before_dedup=[{'h1'}, {'h1'}, {'h2'}])
        r._build_data_section_usage_by_node(1, ws, df, managed_node_type='direct')
        headers = [ws.cell(row=1, column=c).value for c in range(1, 12)]
        assert 'Facts' in headers

    def test_dedup_label_present_when_dedup_enabled(self):
        r = _report(dedup=True)
        ws = _make_ws(r)
        df = _usage_by_node_df(host_names_before_dedup=[{'h1'}, {'h1'}, {'h2'}])
        r._build_data_section_usage_by_node(1, ws, df)
        headers = [ws.cell(row=1, column=c).value for c in range(1, 12)]
        assert any(h and 'deduplication' in h for h in headers)


# ---------------------------------------------------------------------------
# _build_data_section_usage_by_collections
# ---------------------------------------------------------------------------


def _collections_df():
    return pd.DataFrame(
        {
            'collection_name': ['col.a', 'col.b', 'col.a'],
            'host_name': ['h1', 'h2', 'h3'],
            'host_composite_id': ['h1__u1__j1', 'h2__u1__j2', 'h3__u1__j3'],
            'task_runs': [10, 5, 8],
            'duration': [100.0, 50.0, 80.0],
        }
    )


def _indirects_df(**extra):
    # events is the merged, already-parsed list of fully-qualified content names per row
    data = {
        'host_name': ['h1', 'ind_h2'],
        'install_uuid': ['u1', 'u1'],
        'job_remote_id': ['j9', 'j10'],
        'task_runs': [4, 6],
        'events': [['col.a.some_module'], ['col.c.other_module', 'col.a.another_module']],
    }
    data.update(extra)
    return pd.DataFrame(data)


def _sheet_rows(ws, ncols):
    """Read a built worksheet into a list of row tuples, until the first empty first-column cell."""
    rows = []
    r_idx = 1
    while ws.cell(row=r_idx, column=1).value is not None:
        rows.append(tuple(ws.cell(row=r_idx, column=c).value for c in range(1, ncols + 1)))
        r_idx += 1
    return rows


class TestBuildDataSectionUsageByCollections:
    def test_returns_advanced_row_number(self):
        r = _report()
        ws = _make_ws(r)
        end_row = r._build_data_section_usage_by_collections(1, ws, _collections_df())
        assert end_row > 1

    def test_header_is_bold(self):
        r = _report()
        ws = _make_ws(r)
        r._build_data_section_usage_by_collections(1, ws, _collections_df())
        assert ws.cell(row=1, column=1).font.bold is True

    def test_collection_name_in_headers(self):
        r = _report()
        ws = _make_ws(r)
        r._build_data_section_usage_by_collections(1, ws, _collections_df())
        headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
        assert 'Collection name' in headers

    def test_indirect_none_matches_direct_only(self):
        r_direct = _report()
        ws_direct = _make_ws(r_direct)
        r_direct._build_data_section_usage_by_collections(1, ws_direct, _collections_df())

        r_none = _report()
        ws_none = _make_ws(r_none)
        r_none._build_data_section_usage_by_collections(1, ws_none, _collections_df(), indirects=None)

        assert _sheet_rows(ws_direct, 5) == _sheet_rows(ws_none, 5)

    def test_indirect_empty_adds_no_collections(self):
        r = _report()
        ws = _make_ws(r)
        r._build_data_section_usage_by_collections(1, ws, _collections_df(), indirects=_indirects_df().iloc[0:0])
        collection_names = {row[0] for row in _sheet_rows(ws, 5)[1:]}
        assert collection_names == {'col.a', 'col.b'}

    def test_indirect_merges_and_adds_collections(self):
        r = _report()
        ws = _make_ws(r)
        r._build_data_section_usage_by_collections(1, ws, _collections_df(), indirects=_indirects_df())

        by_collection = {row[0]: row for row in _sheet_rows(ws, 5)[1:]}

        # col.a gains indirect h1 (same name, new composite id) and ind_h2: unique host
        # names {h1, h3, ind_h2}=3, non-unique composite ids=4, task_runs 10+8+4+6=28
        assert by_collection['col.a'] == ('col.a', 3, 4, 28, 180.0)
        assert by_collection['col.b'] == ('col.b', 1, 1, 5, 50.0)
        # col.c is indirect-only (ind_h2), no duration recorded for indirect
        assert by_collection['col.c'] == ('col.c', 1, 1, 6, 0.0)


class TestBuildIndirectCollectionsLong:
    def test_returns_none_when_indirects_none(self):
        assert _report()._build_indirect_collections_long(None) is None

    def test_returns_none_when_indirects_empty(self):
        assert _report()._build_indirect_collections_long(_indirects_df().iloc[0:0]) is None

    def test_returns_none_when_no_events_resolve_to_collections(self):
        df = _indirects_df(events=[['not_a_fqcn'], []])
        assert _report()._build_indirect_collections_long(df) is None

    def test_returns_none_when_events_empty(self):
        df = _indirects_df(events=[[], None])
        assert _report()._build_indirect_collections_long(df) is None

    def test_explodes_events_into_one_row_per_collection(self):
        result = _report()._build_indirect_collections_long(_indirects_df())
        # h1 -> col.a ; ind_h2 -> col.c, col.a
        assert len(result) == 3
        assert set(result['collection_name']) == {'col.a', 'col.c'}

    def test_computes_composite_id_and_zero_duration(self):
        df = _indirects_df(events=[['col.a.mod'], []])
        result = _report()._build_indirect_collections_long(df)
        assert list(result['host_composite_id']) == ['h1__u1__j9']
        assert list(result['duration']) == [0]

    def test_task_runs_attributed_to_each_collection_in_row(self):
        df = _indirects_df(host_name=['ind_h2'], install_uuid=['u1'], job_remote_id=['j10'], task_runs=[6], events=[['col.a.m1', 'col.c.m2']])
        result = _report()._build_indirect_collections_long(df)
        assert sorted(zip(result['collection_name'], result['task_runs'])) == [('col.a', 6), ('col.c', 6)]

    def test_output_columns_match_long_schema(self):
        result = _report()._build_indirect_collections_long(_indirects_df())
        assert list(result.columns) == Base.COLLECTIONS_LONG_COLUMNS


# ---------------------------------------------------------------------------
# _build_data_section_usage_by_roles
# ---------------------------------------------------------------------------


def _roles_df():
    return pd.DataFrame(
        {
            'role_name': ['role.x', 'role.y', 'role.x'],
            'host_name': ['h1', 'h2', 'h3'],
            'host_composite_id': ['h1__u1__j1', 'h2__u1__j2', 'h3__u1__j3'],
            'task_runs': [4, 6, 2],
            'duration': [40.0, 60.0, 20.0],
        }
    )


class TestBuildDataSectionUsageByRoles:
    def test_returns_advanced_row_number(self):
        r = _report()
        ws = _make_ws(r)
        end_row = r._build_data_section_usage_by_roles(1, ws, _roles_df())
        assert end_row > 1

    def test_header_is_bold(self):
        r = _report()
        ws = _make_ws(r)
        r._build_data_section_usage_by_roles(1, ws, _roles_df())
        assert ws.cell(row=1, column=1).font.bold is True

    def test_role_name_in_headers(self):
        r = _report()
        ws = _make_ws(r)
        r._build_data_section_usage_by_roles(1, ws, _roles_df())
        headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
        assert 'Role name' in headers


# ---------------------------------------------------------------------------
# _build_data_section_usage_by_modules
# ---------------------------------------------------------------------------


def _modules_df():
    return pd.DataFrame(
        {
            'module_name': ['mod.a', 'mod.b', 'mod.a'],
            'host_name': ['h1', 'h2', 'h3'],
            'host_composite_id': ['h1__u1__j1', 'h2__u1__j2', 'h3__u1__j3'],
            'task_runs': [3, 9, 1],
            'duration': [30.0, 90.0, 10.0],
        }
    )


class TestBuildDataSectionUsageByModules:
    def test_returns_advanced_row_number(self):
        r = _report()
        ws = _make_ws(r)
        end_row = r._build_data_section_usage_by_modules(1, ws, _modules_df())
        assert end_row > 1

    def test_header_is_bold(self):
        r = _report()
        ws = _make_ws(r)
        r._build_data_section_usage_by_modules(1, ws, _modules_df())
        assert ws.cell(row=1, column=1).font.bold is True

    def test_module_name_in_headers(self):
        r = _report()
        ws = _make_ws(r)
        r._build_data_section_usage_by_modules(1, ws, _modules_df())
        headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
        assert 'Module name' in headers
