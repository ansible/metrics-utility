from datetime import datetime
import os
import subprocess
import sys

import openpyxl
import pytest

env_vars = {
    "METRICS_UTILITY_PRICE_PER_NODE": "11.55",
    "METRICS_UTILITY_REPORT_RHN_LOGIN": "test_login",
    "METRICS_UTILITY_SHIP_PATH": "/awx_devel/awx-dev/metrics-utility/metrics_utility/test/test_data",
    "METRICS_UTILITY_REPORT_SKU_DESCRIPTION": "EX: Red Hat Ansible Automation Platform, Full Support (1 Managed Node, Dedicated, Monthly)",
    "METRICS_UTILITY_REPORT_H1_HEADING": "CCSP NA Direct Reporting Template",
    "METRICS_UTILITY_REPORT_PO_NUMBER": "123",
    "METRICS_UTILITY_SHIP_TARGET": "directory",
    "METRICS_UTILITY_REPORT_COMPANY_NAME": "Partner A",
    "METRICS_UTILITY_REPORT_SKU": "MCT3752MO",
    "METRICS_UTILITY_REPORT_EMAIL": "email@email.com",
    "METRICS_UTILITY_REPORT_TYPE": "CCSP",
    "AWX_LOGGING_MODE": "stdout",
}

file_path = "/awx_devel/awx-dev/metrics-utility/metrics_utility/test/test_data/reports/2024/02/CCSP-2024-02.xlsx"


date_today = datetime.now().strftime("%b %d, %Y")

EXPECTED_SHEETS = {
    "Usage Reporting": [
        {
            "SKU": [
                'CCSP Company Name',
                'CCSP Email',
                'CCSP RHN Login',
                'Report Period (YYYY-MM)',
                'Company Business leader ',
                'Company Procurement leader ',
                'Periodicity', None,
                'Monthly payment',
                None,
                None,
                None,
                'SKU',
                'MCT3752MO',
                None,
                'Organization name (i.e. company name)',
                'Default',
                'test organization',
                'Sum of monthly payment']
        },
        {
            "SKU Description": [
                'Partner A',
                'email@email.com',
                'test_login',
                '2024-02',
                None,
                None,
                'Report is submitted monthly',
                None,
                None,
                None,
                None,
                None,
                'SKU Description',
                'EX: Red Hat Ansible Automation Platform, Full Support (1 Managed Node, Dedicated, Monthly)',
                None,
                "Please Mark With An 'X' If The Usage Is Internal. \nOtherwise Leave Blank", None, None, None]
        },
        {
            '': [
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               'Red Hat SKU\n Quantity Consumed',
               1,
               2,
               None,
            ]
        },
        {
            "Term": [
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               'Term',
               'MONTH',
               None,
               'Subscription Fee\n (SKU Unit Price)',
               11.55,
               11.55,
               None,
            ]
        },
        {
            "Unit of Measure": [
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               'Unit of Measure',
               'MANAGED NODE',
               None,
               'Extended\n Subscription Fees\n (SKU Extended Unit Price)',
               '=C18*D18',
               '=C19*D19',
               '=SUM(E18:E19)',
            ]
        },
         {
            "Currency": [
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               'Currency',
               'USD',
               None,
               None,
               None,
               None,
               None,
            ]
        },
         {
            "MSRP": [
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               None,
               'MSRP',
               '11.55',
               None,
               None,
               None,
               None,
               None,
            ]
        },
    ],
    "Managed nodes": [
        {"Host name": ['localhost', 'test host 1', 'test host 2']},
        {"automated by organizations": [1, 1, 1]},
        {'job runs': [2, 2, 2]},
        {'number of task runs': [4, 4, 4]},
        {
            'first automation': [
                datetime(2024, 2, 28, 8, 48, 36, 37000),
                datetime(2024, 2, 28, 8, 48, 41, 638000),
                datetime(2024, 2, 28, 8, 48, 41, 638000),
            ]
        },
        {
            'last automation': [
                datetime(2024, 2, 28, 8, 48, 50, 35000),
                datetime(2024, 2, 28, 8, 48, 58, 766000),
                datetime(2024, 2, 28, 8, 48, 58, 766000)
            ]
        }
    ]
}



@pytest.fixture
def cleanup():
    """Fixture to clean up the generated file at the start and end of test."""
    # Cleanup at the beginning
    if os.path.exists(file_path):
        os.remove(file_path)
    yield
    # Cleanup at the end
    if os.path.exists(file_path):
        os.remove(file_path)


def validate_sheet_tab_names():
    """Test the sheet names in the Excel file."""

    wb = openpyxl.load_workbook(file_path)
    try:
        actual_tab_names = wb.sheetnames
        assert actual_tab_names == list(
            EXPECTED_SHEETS.keys()
        ), "Sheet names do not match."
    finally:
        wb.close()


def validate_sheet_columns():
    """Test the column names for each sheet."""

    def normalize_column(col):
        return col.strip().replace("\n", " ").lower() if col else ""

    wb = openpyxl.load_workbook(file_path)
    try:
        for sheet_name, expected_column_data in EXPECTED_SHEETS.items():
            sheet = wb[sheet_name]

            # For the 'Usage Reporting' sheet, start at row 14
            if sheet_name == "Usage Reporting":
                min_row = 14
            else:
                min_row = 1  # Default for other sheets

            # All actual column headers for sheet
            actual_column_headers = [normalize_column(cell.value) for cell in next(sheet.iter_rows(min_row=min_row, max_row=min_row))]

            # All expected column headers
            expected_column_headers = []
            for column_group in expected_column_data:
                expected_column_headers.extend(normalize_column(col) for col in column_group.keys())

            print("Actual column headers (formatted):", actual_column_headers)
            print("Expected column headers (formatted):", expected_column_headers)

            # Assert column headers
            assert actual_column_headers == expected_column_headers, f"Column names do not match for sheet: {sheet_name}"

            # Iterate through each expected column group
            for column_group in expected_column_data:
                for expected_col_name, expected_column_values in column_group.items():

                    # Find the actual column index for this column
                    try:
                        col_index = actual_column_headers.index(normalize_column(expected_col_name)) + 1
                    except ValueError:
                        raise AssertionError(f"Expected column '{expected_col_name}' not found in actual columns for sheet: {sheet_name}")

                    # Extract actual values for this column (skip the header)
                    actual_column_values = [
                        cell.value for row in sheet.iter_rows(min_row=2, min_col=col_index, max_col=col_index)
                        for cell in row
                    ]

                    print(f"Actual column values for '{expected_col_name}':", actual_column_values)
                    print(f"Expected column values for '{expected_col_name}':", expected_column_values)

                    # Assert column values
                    assert actual_column_values == expected_column_values, (
                        f"Column values do not match for column '{expected_col_name}' in sheet '{sheet_name}'"
                    )
    finally:
        wb.close()

def test_command(cleanup):
    """Build xlsx report using build command and test its contents."""

    python_executable = sys.executable
    result = subprocess.run(
        [python_executable, "manage.py", "build_report", "--month=2024-02", "--force"],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        env=env_vars,
    )

    assert result.returncode == 0

    validate_sheet_columns()
    validate_sheet_tab_names()
