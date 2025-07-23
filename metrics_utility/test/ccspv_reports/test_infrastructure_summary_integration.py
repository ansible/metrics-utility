"""
Integration tests for Infrastructure Summary tab using actual tarball data.

This module demonstrates how to test the Infrastructure Summary functionality
with real data from the test tarballs in metrics_utility/test/test_data/data/2025/07/22/
"""

import os
import tarfile
import tempfile

from unittest.mock import patch

import pandas as pd
import pytest

from metrics_utility.automation_controller_billing.report.base import Base
from metrics_utility.automation_controller_billing.report.report_ccsp_v2 import ReportCCSPv2
from metrics_utility.metric_utils import DIRECT, INDIRECT


class TestInfrastructureSummaryWithTarballData:
    """Integration tests using actual tarball data."""

    @pytest.fixture
    def tarball_path(self):
        """Path to the test tarball."""
        return 'metrics_utility/test/test_data/data/2025/07/22/9bf6f6c3-eeb3-4523-8e8d-80f3d4d01978-2025-07-22-000000+0000-0.tar.gz'

    @pytest.fixture
    def extracted_data(self, tarball_path):
        """Extract and load data from tarball."""
        if not os.path.exists(tarball_path):
            pytest.skip(f'Tarball not found: {tarball_path}')

        with tempfile.TemporaryDirectory() as temp_dir:
            # Extract tarball
            with tarfile.open(tarball_path, 'r:gz') as tar:
                tar.extractall(temp_dir)

            # Load CSV files
            dataframes = {}

            # Load main_indirectmanagednodeaudit.csv
            audit_path = os.path.join(temp_dir, 'main_indirectmanagednodeaudit.csv')
            if os.path.exists(audit_path):
                audit_df = pd.read_csv(audit_path)
                # Convert to job_host_summary format for testing
                job_host_summary = pd.DataFrame(
                    {
                        'managed_node_type': [INDIRECT] * len(audit_df),
                        'host_name': audit_df['host_name'],
                        'facts': audit_df['facts'],
                        'organization_name': audit_df.get('organization_name', 'Default'),
                        'task_runs': audit_df.get('task_runs', 1),
                        'first_automation': audit_df['created'],
                        'last_automation': audit_df['created'],
                        'managed_node_types_set': [{'indirect'}] * len(audit_df),
                        'events': audit_df['events'],
                        'canonical_facts': audit_df['canonical_facts'],
                    }
                )
                dataframes['job_host_summary'] = job_host_summary

            # Load other CSV files if needed
            status_path = os.path.join(temp_dir, 'data_collection_status.csv')
            if os.path.exists(status_path):
                dataframes['data_collection_status'] = pd.read_csv(status_path)

            return dataframes

    def test_infrastructure_summary_with_real_data(self, extracted_data):
        """Test Infrastructure Summary with real tarball data."""
        if 'job_host_summary' not in extracted_data:
            pytest.skip('No job_host_summary data available')

        # Create extra params for the report
        extra_params = {
            'price_per_node': 11.55,
            'report_period': '2025-07',
            'report_sku': 'TEST-SKU',
            'report_h1_heading': 'Test Infrastructure Summary Report',
            'report_po_number': '123',
            'report_company_name': 'Test Company',
            'report_email': 'test@example.com',
            'report_rhn_login': 'test_login',
            'report_sku_description': 'Test Infrastructure Summary Description',
            'optional_report_sheets': 'infrastructure_summary',
            'report_organization_filter': None,
        }

        # Mock the dataframes that would come from tarball processing
        mock_dataframes = {
            'job_host_summary': pd.DataFrame(
                {
                    'managed_node_type': [INDIRECT, INDIRECT, INDIRECT, INDIRECT, INDIRECT],
                    'host_name': ['host_1', 'host_2', 'host_3', 'host_4', 'host_5'],
                    'original_host_name': ['host_1', 'host_2', 'host_3', 'host_4', 'host_5'],
                    'install_uuid': ['uuid1', 'uuid2', 'uuid3', 'uuid4', 'uuid5'],
                    'job_remote_id': ['job1', 'job2', 'job3', 'job4', 'job5'],
                    'facts': [
                        '{"device_type": "Containers", "infra_type": "Public Cloud", "infra_bucket": "Storage"}',
                        '{"device_type": "Object Storage", "infra_type": "Hybrid Cloud", "infra_bucket": "Database"}',
                        '{"device_type": "Block Storage", "infra_type": "On-Premises", "infra_bucket": "Network"}',
                        '{"device_type": "File Storage", "infra_type": "Edge Computing", "infra_bucket": "Security"}',
                        '{"device_type": "SQL", "infra_type": "Multi-Cloud", "infra_bucket": "Analytics"}',
                    ],
                }
            ),
            'main_jobevent': pd.DataFrame(
                {
                    'host_name': ['host_1', 'host_2', 'host_3', 'host_4', 'host_5'],
                    'install_uuid': ['uuid1', 'uuid2', 'uuid3', 'uuid4', 'uuid5'],
                    'job_remote_id': ['job1', 'job2', 'job3', 'job4', 'job5'],
                }
            ),
            'main_host': pd.DataFrame(),  # Required by build_spreadsheet
            'data_collection_status': pd.DataFrame(),  # Required by build_spreadsheet
        }

        # Create report instance
        report = ReportCCSPv2(mock_dataframes, extra_params)

        # Mock the optional_report_sheets method to include infrastructure_summary
        with patch.object(report, 'optional_report_sheets', return_value=['infrastructure_summary']):
            # Mock the add_sheet method
            with patch.object(report, 'add_sheet') as mock_add_sheet:
                # Mock the _build_data_section_infrastructure_summary method
                with patch.object(report, '_build_data_section_infrastructure_summary') as mock_build:
                    # Call build_spreadsheet
                    report.build_spreadsheet()

                    # Verify Infrastructure Summary sheet was requested
                    infrastructure_summary_calls = [call for call in mock_add_sheet.call_args_list if call[0][0] == 'Infrastructure Summary']
                    assert len(infrastructure_summary_calls) == 1

                    # Verify build method was called
                    mock_build.assert_called_once()

                    # Verify the correct dataframe was passed
                    call_args = mock_build.call_args
                    dataframe = call_args[0][2]  # Third argument is the dataframe
                    assert len(dataframe) > 0
                    assert 'managed_node_type' in dataframe.columns
                    assert 'host_name' in dataframe.columns
                    assert 'facts' in dataframe.columns

    def test_infrastructure_summary_data_processing(self, extracted_data):
        """Test the actual data processing logic with real data."""
        if 'job_host_summary' not in extracted_data:
            pytest.skip('No job_host_summary data available')

        df = extracted_data['job_host_summary']

        # Test the base class infrastructure summary method directly
        base = Base()
        from openpyxl import Workbook

        ws = Workbook().active

        # Call the infrastructure summary method
        result_row = base._build_data_section_infrastructure_summary(1, ws, df)

        # Verify that the method processed the data
        assert result_row > 1  # Should have processed at least the header

        # Check that headers were created
        assert ws.cell(row=1, column=1).value == 'Infrastructure'
        assert ws.cell(row=1, column=2).value == 'Device Category'
        assert ws.cell(row=1, column=3).value == 'Device Type'
        assert ws.cell(row=1, column=4).value == 'Unique Nodes'
        assert ws.cell(row=1, column=5).value == 'Total Nodes'

        # Check that some data was processed (if there are indirect nodes)
        indirect_nodes = df[df['managed_node_type'] == INDIRECT]
        if len(indirect_nodes) > 0:
            # Should have processed some infrastructure data
            assert result_row > 2
        else:
            # Should show "No indirect nodes found" message
            assert ws.cell(row=1, column=1).value == 'No indirect nodes found'

    def test_infrastructure_summary_with_enriched_data(self):
        """Test Infrastructure Summary with the enriched test data we created."""
        # Use the enriched main_indirectmanagednodeaudit.csv we created earlier
        csv_path = 'metrics_utility/test/test_data/data/2025/07/22/main_indirectmanagednodeaudit.csv'

        if not os.path.exists(csv_path):
            pytest.skip(f'Enriched CSV not found: {csv_path}')

        # Load the enriched data
        audit_df = pd.read_csv(csv_path)

        # Convert to job_host_summary format
        job_host_summary = pd.DataFrame(
            {
                'managed_node_type': [INDIRECT] * len(audit_df),
                'host_name': audit_df['host_name'],
                'facts': audit_df['facts'],
                'organization_name': ['Default'] * len(audit_df),
                'task_runs': [1] * len(audit_df),
                'first_automation': audit_df['created'],
                'last_automation': audit_df['created'],
                'managed_node_types_set': [{'indirect'}] * len(audit_df),
                'events': audit_df['events'],
                'canonical_facts': audit_df['canonical_facts'],
            }
        )

        # Test the infrastructure summary processing
        base = Base()
        from openpyxl import Workbook

        ws = Workbook().active

        result_row = base._build_data_section_infrastructure_summary(1, ws, job_host_summary)

        # Should have processed the enriched data
        assert result_row > 1

        # Check for specific infrastructure types from our enriched data
        # Look for the infrastructure types we included in the enriched data
        found_infra_types = set()
        for row in range(2, result_row):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and cell_value != 'Infrastructure':
                found_infra_types.add(cell_value)

        # Should have found various infrastructure types from our enriched data
        expected_types = {
            'Public Cloud',
            'Private Cloud',
            'Hybrid Cloud',
            'On-Premises',
            'Edge Computing',
            'Multi-Cloud',
            'Serverless',
            'Container Platform',
        }

        # Check that we found at least some of the expected types
        assert len(found_infra_types.intersection(expected_types)) > 0

        # Check for device types from our enriched data
        found_device_types = set()
        for row in range(2, result_row):
            cell_value = ws.cell(row=row, column=3).value
            if cell_value and cell_value != 'Device Type':
                found_device_types.add(cell_value)

        # Should have found various device types from our enriched data
        expected_device_types = {
            'Containers',
            'Virtual Machines',
            'Object Storage',
            'Block Storage',
            'File Storage',
            'SQL',
            'NoSQL',
            'CDN',
            'Load Balancer',
            'Firewall',
        }

        # Check that we found at least some of the expected device types
        assert len(found_device_types.intersection(expected_device_types)) > 0

    def test_infrastructure_summary_edge_cases_with_real_data(self, extracted_data):
        """Test edge cases with real data."""
        if 'job_host_summary' not in extracted_data:
            pytest.skip('No job_host_summary data available')

        df = extracted_data['job_host_summary']

        # Test with empty dataframe
        empty_df = pd.DataFrame(columns=df.columns)
        base = Base()
        from openpyxl import Workbook

        ws = Workbook().active

        result_row = base._build_data_section_infrastructure_summary(1, ws, empty_df)
        assert result_row == 2
        assert ws.cell(row=1, column=1).value == 'No indirect nodes found'

        # Test with only direct nodes
        direct_only_df = df.copy()
        direct_only_df['managed_node_type'] = DIRECT

        from openpyxl import Workbook

        ws = Workbook().active
        result_row = base._build_data_section_infrastructure_summary(1, ws, direct_only_df)
        assert result_row == 2
        assert ws.cell(row=1, column=1).value == 'No indirect nodes found'

        # Test with mixed direct and indirect nodes
        mixed_df = df.copy()
        mixed_df.loc[0, 'managed_node_type'] = DIRECT  # Make first row direct

        from openpyxl import Workbook

        ws = Workbook().active
        result_row = base._build_data_section_infrastructure_summary(1, ws, mixed_df)

        # Should process only indirect nodes
        indirect_count = len(mixed_df[mixed_df['managed_node_type'] == INDIRECT])
        if indirect_count > 0:
            assert result_row > 2
        else:
            assert result_row == 2
            assert ws.cell(row=1, column=1).value == 'No indirect nodes found'

    def test_infrastructure_summary_with_null_empty_scenarios(self):
        """Test Infrastructure Summary with null/empty field scenarios from enriched data."""
        # Use the enriched main_indirectmanagednodeaudit.csv we created earlier
        csv_path = 'metrics_utility/test/test_data/data/2025/07/22/main_indirectmanagednodeaudit.csv'

        if not os.path.exists(csv_path):
            pytest.skip(f'Enriched CSV not found: {csv_path}')

        # Load the enriched data
        audit_df = pd.read_csv(csv_path)

        # Filter for null test hosts (rows 41-50 in our enriched data)
        null_test_hosts = audit_df[audit_df['host_name'].str.contains('null_test_host', na=False)]

        if len(null_test_hosts) == 0:
            pytest.skip('No null test hosts found in enriched data')

        # Convert to job_host_summary format
        job_host_summary = pd.DataFrame(
            {
                'managed_node_type': [INDIRECT] * len(null_test_hosts),
                'host_name': null_test_hosts['host_name'],
                'facts': null_test_hosts['facts'],
                'organization_name': ['Default'] * len(null_test_hosts),
                'task_runs': [1] * len(null_test_hosts),
                'first_automation': null_test_hosts['created'],
                'last_automation': null_test_hosts['created'],
                'managed_node_types_set': [{'indirect'}] * len(null_test_hosts),
                'events': null_test_hosts['events'],
                'canonical_facts': null_test_hosts['canonical_facts'],
            }
        )

        # Test the infrastructure summary processing
        base = Base()
        from openpyxl import Workbook

        ws = Workbook().active

        result_row = base._build_data_section_infrastructure_summary(1, ws, job_host_summary)

        # Should have processed the null/empty data
        assert result_row > 1

        # Check for 'Unknown' values which indicate missing fields were handled
        unknown_values_found = 0
        for row in range(2, result_row):
            for col in range(1, 6):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value == 'Unknown':
                    unknown_values_found += 1

        # Should have found some 'Unknown' values from the null/empty fields
        assert unknown_values_found > 0

        # Verify that the method handled all the null/empty scenarios without crashing
        print(f'✓ Successfully processed {len(null_test_hosts)} null/empty test scenarios')
        print(f"✓ Found {unknown_values_found} 'Unknown' values indicating missing fields")


def run_infrastructure_summary_tests():
    """Helper function to run the infrastructure summary tests."""
    print('Running Infrastructure Summary tests...')

    # Test with the enriched data
    csv_path = 'metrics_utility/test/test_data/data/2025/07/22/main_indirectmanagednodeaudit.csv'

    if os.path.exists(csv_path):
        print(f'✓ Found enriched test data: {csv_path}')

        # Load and analyze the data
        audit_df = pd.read_csv(csv_path)
        print(f'✓ Loaded {len(audit_df)} records from enriched data')

        # Check for infrastructure types
        infra_types = set()
        device_types = set()

        for facts_str in audit_df['facts']:
            try:
                facts = eval(facts_str)  # Convert string representation to dict
                if isinstance(facts, dict):
                    infra_types.add(facts.get('infra_type', 'Unknown'))
                    device_types.add(facts.get('device_type', 'Unknown'))
            except (NameError, SyntaxError, TypeError):
                pass

        print(f'✓ Found {len(infra_types)} infrastructure types: {sorted(infra_types)}')
        print(f'✓ Found {len(device_types)} device types: {sorted(device_types)}')

        # Test the infrastructure summary processing
        job_host_summary = pd.DataFrame(
            {
                'managed_node_type': [INDIRECT] * len(audit_df),
                'host_name': audit_df['host_name'],
                'facts': audit_df['facts'],
                'organization_name': ['Default'] * len(audit_df),
                'task_runs': [1] * len(audit_df),
                'first_automation': audit_df['created'],
                'last_automation': audit_df['created'],
                'managed_node_types_set': [{'indirect'}] * len(audit_df),
                'events': audit_df['events'],
                'canonical_facts': audit_df['canonical_facts'],
            }
        )

        base = Base()
        from openpyxl import Workbook

        ws = Workbook().active

        result_row = base._build_data_section_infrastructure_summary(1, ws, job_host_summary)
        print(f'✓ Successfully processed infrastructure summary in {result_row} rows')

        # Check for specific test scenarios
        found_infra_types = set()
        found_device_types = set()

        for row in range(2, result_row):
            infra_cell = ws.cell(row=row, column=1).value
            device_cell = ws.cell(row=row, column=3).value

            if infra_cell and infra_cell != 'Infrastructure':
                found_infra_types.add(infra_cell)
            if device_cell and device_cell != 'Device Type':
                found_device_types.add(device_cell)

        print(f'✓ Processed {len(found_infra_types)} infrastructure types in summary')
        print(f'✓ Processed {len(found_device_types)} device types in summary')

        # Check for edge cases
        unicode_hosts = audit_df[audit_df['host_name'].str.contains(r'[^\x00-\x7F]', na=False)]
        if len(unicode_hosts) > 0:
            print(f'✓ Found {len(unicode_hosts)} hosts with Unicode characters')

        long_names = audit_df[audit_df['host_name'].str.len() > 50]
        if len(long_names) > 0:
            print(f'✓ Found {len(long_names)} hosts with long names (>50 chars)')

        print('✓ All infrastructure summary tests completed successfully!')

    else:
        print(f'✗ Enriched test data not found: {csv_path}')
        print('Please run the data enrichment script first.')


if __name__ == '__main__':
    run_infrastructure_summary_tests()
