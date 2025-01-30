import os
import subprocess
import sys
from datetime import datetime

import openpyxl
import pytest

env_vars = {
    "METRICS_UTILITY_PRICE_PER_NODE": "11.55",
    "METRICS_UTILITY_REPORT_RHN_LOGIN": "test_login",
    "METRICS_UTILITY_SHIP_PATH": "/awx_devel/awx-dev/metrics-utility/metrics_utility/test/test_data",
    "METRICS_UTILITY_REPORT_END_USER_COMPANY_NAME": "Customer A",
    "METRICS_UTILITY_REPORT_END_USER_STATE": "TX",
    "METRICS_UTILITY_REPORT_SKU_DESCRIPTION": "EX: Red Hat Ansible Automation Platform, Full Support (1 Managed Node, Dedicated, Monthly)",
    "METRICS_UTILITY_REPORT_H1_HEADING": "CCSP NA Direct Reporting Template",
    "METRICS_UTILITY_REPORT_END_USER_CITY": "Springfield",
    "METRICS_UTILITY_REPORT_PO_NUMBER": "123",
    "METRICS_UTILITY_SHIP_TARGET": "directory",
    "METRICS_UTILITY_REPORT_END_USER_COUNTRY": "US",
    "METRICS_UTILITY_REPORT_COMPANY_NAME": "Partner A",
    "METRICS_UTILITY_REPORT_SKU": "MCT3752MO",
    "METRICS_UTILITY_REPORT_EMAIL": "email@email.com",
    "METRICS_UTILITY_REPORT_TYPE": "CCSPv2",
    "AWX_LOGGING_MODE": "stdout",
}

file_path = "/awx_devel/awx-dev/metrics-utility/metrics_utility/test/test_data/reports/2024/02/CCSPv2-2024-02.xlsx"


date_today = datetime.now().strftime("%b %d, %Y")
EXPECTED_SHEETS = {
    "Usage Reporting": [
        "CCSP NA Direct Reporting Template",
        "",
        "",
        "",
        "",
        "",
        "",
        f"Updated: {date_today}",
        "",
        "",
        "",
    ],
    "Managed nodes": [
        {"Host name": ['Host name', 'localhost', 'test host 1', 'test host 2']},
        "Automated by\norganizations",
        "Job runs",
        "Number of task\nruns",
        "First\nautomation",
        "Last\nautomation",
    ],
    "Usage by organizations": [
        {"Organization name" : ['Organization name', 'Default', 'test organization']},
        "Job runs",
        "Unique managed nodes\nautomated",
        "Non-unique managed\nnodes automated",
        "Number of task\nruns",
    ],
}


@pytest.fixture
def cleanup():
    """Fixture to clean up the generated file at the start and end of test."""
    # Cleanup at the beginning
    if os.path.exists(file_path):
        os.remove(file_path)
    yield
    # Cleanup at the end
    if os.path.exists(file_path):
        os.remove(file_path)


def validate_sheet_tab_names():
    """Test the sheet names in the Excel file."""

    wb = openpyxl.load_workbook(file_path)
    try:
        actual_tab_names = wb.sheetnames
        assert actual_tab_names == list(
            EXPECTED_SHEETS.keys()
        ), "Sheet names do not match."
    finally:
        wb.close()


def validate_sheet_columns():
    """Test the column names for each sheet."""

    def normalize_column(col):
        return col.strip().replace("\n", " ").lower() if col else ""

    wb = openpyxl.load_workbook(file_path)
    try:
        for sheet_name, expected_column_data in EXPECTED_SHEETS.items():
            sheet = wb[sheet_name]
            actual_column_headers = [normalize_column(cell.value) for cell in next(sheet.iter_rows(max_row=1))]

            for column_group in expected_column_data:
                expected_column_headers = [normalize_column(col) for col in list(column_group.keys())]
            if actual_column_headers != expected_column_headers:
                print(f"Mismatch for sheet: {sheet_name}")
                print(f"Actual columns (formatted): {actual_column_headers}")
                print(f"Expected columns (formatted): {expected_column_headers}")

            # check column headers
            assert (
                actual_column_headers == expected_column_headers
            ), f"Column names do not match for sheet: {sheet_name}"

            # check column values
            # expected
    finally:
            wb.close()

def check_numeric_values(file_path, sheet_name, column_index):
    """
    Checks if all the values in a specified column are numbers.

    :param file_path: Path to the Excel file.
    :param sheet_name: Name of the sheet to validate.
    :param column_index: Index of the column to check (1-based, e.g., 1 for 'A').
    :return: True if all values in the column are numbers, False otherwise.
    """
    wb = openpyxl.load_workbook(file_path)
    try:
        for sheet_name, expected_columns in EXPECTED_SHEETS.items():
            sheet = wb[sheet_name]
            print(expected_columns)

            actual_values = [row[0].value for row in sheet.iter_rows(min_col=1, max_col=1, values_only=False)]
            print(actual_values)
    finally:
        wb.close()

def test_command():
    """Build xlsx report using build command and test its contents."""

    python_executable = sys.executable
    result = subprocess.run(
        [python_executable, "manage.py", "build_report", "--month=2024-02", "--force"],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        env=env_vars,
    )

    assert result.returncode == 0

    validate_sheet_columns()
    validate_sheet_tab_names()
