import os
import pytest
import openpyxl

class Helpers:
    @staticmethod
    def validate_sheet_tab_names(file_path, EXPECTED_SHEETS):
        """Test the sheet names in the Excel file."""

        wb = openpyxl.load_workbook(file_path)
        try:
            actual_tab_names = wb.sheetnames
            assert actual_tab_names == list(
                EXPECTED_SHEETS.keys()
            ), "Sheet names do not match."
        finally:
            wb.close()

    @staticmethod
    def validate_sheet_columns(file_path, EXPECTED_SHEETS, usage_reporting_min_row):
        """Test the column names for each sheet."""

        def normalize_column(col):
            return col.strip().replace("\n", " ").lower() if col else ""

        wb = openpyxl.load_workbook(file_path)
        try:
            for sheet_name, expected_column_data in EXPECTED_SHEETS.items():
                sheet = wb[sheet_name]

                # For the 'Usage Reporting' sheet, start at specific row with value for ease of traversing
                if sheet_name == "Usage Reporting":
                    min_row = usage_reporting_min_row
                else:
                    min_row = 1  # Default for other sheets

                # All actual column headers for sheet
                actual_column_headers = [normalize_column(cell.value) for cell in next(sheet.iter_rows(min_row=min_row, max_row=min_row))]

                # All expected column headers
                expected_column_headers = []
                for column_group in expected_column_data:
                    expected_column_headers.extend(normalize_column(col) for col in column_group.keys())

                print("Actual column headers (formatted):", actual_column_headers)
                print("Expected column headers (formatted):", expected_column_headers)

                # Assert column headers
                assert actual_column_headers == expected_column_headers, f"Column names do not match for sheet: {sheet_name}"

                # Iterate through each expected column group
                for column_group in expected_column_data:
                    for expected_col_name, expected_column_values in column_group.items():

                        # Find the actual column index for this column
                        try:
                            col_index = actual_column_headers.index(normalize_column(expected_col_name)) + 1
                        except ValueError:
                            raise AssertionError(f"Expected column '{expected_col_name}' not found in actual columns for sheet: {sheet_name}")

                        # Extract actual values for this column (skip the header)
                        actual_column_values = [
                            cell.value for row in sheet.iter_rows(min_row=2, min_col=col_index, max_col=col_index)
                            for cell in row
                        ]

                        print(f"Actual column values for '{expected_col_name}':", actual_column_values)
                        print(f"Expected column values for '{expected_col_name}':", expected_column_values)

                        # Assert column values
                        assert actual_column_values == expected_column_values, (
                            f"Column values do not match for column '{expected_col_name}' in sheet '{sheet_name}'"
                        )
        finally:
            wb.close()



@pytest.fixture
def cleanup(request):
    file_path = request.param
    """Fixture to clean up the generated file at the start and end of test."""
    # Cleanup at the beginning
    if os.path.exists(file_path):
        os.remove(file_path)
    yield
    # Cleanup at the end
    if os.path.exists(file_path):
        os.remove(file_path)
