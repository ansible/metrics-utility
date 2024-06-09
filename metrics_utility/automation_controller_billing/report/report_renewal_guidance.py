######################################
# Code for building the spreadsheet
######################################
from metrics_utility.automation_controller_billing.helpers import parse_number_of_days
from metrics_utility.automation_controller_billing.report.base import Base
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

import datetime
import pandas as pd
import time

class ReportRenewalGuidance(Base):
    def __init__(self, dataframe, report_period, extra_params):
        self.wb = Workbook()

        self.dataframe = dataframe
        self.report_period = report_period
        self.extra_params = extra_params
        self.price_per_node = extra_params['price_per_node']

        self.config = {
            'h1_heading': {
                'value': "Renewal guidance",
            },
            'header': [{
                'label': 'Report Period (YYYY-MM-DD, YYYY-MM-DD)',
                'value': '<autogenerated>',
            }],
        }

        default_column_widths = {
            1: 45,
            2: 30,
            3: 15,
            4: 15,
            5: 15,
            6: 20,
            7: 15,
            8: 30,
            9: 20,
            10: 20,
            11: 20
        }

        default_data_column_widths = {
            1: 40,
            2: 20,
            3: 20,
            4: 20,
            5: 20,
            6: 20,
            7: 20
        }

        uniform_column_widths = {
            1: 20,
            2: 20,
            3: 20,
            4: 20,
            5: 20,
            6: 20,
            7: 20
        }

        self.config['column_widths'] = default_column_widths
        self.config['data_column_widths'] = default_data_column_widths
        self.config['uniform_column_widths'] = uniform_column_widths


    def build_spreadsheet(self):
        # Fix host names in the event data, to take in account the variables
        host_metric_dataframe = self.dataframe[0]
        # Spreadsheet doesn't support timezones
        host_metric_dataframe['first_automation'] = pd.to_datetime(
            host_metric_dataframe['first_automation']).dt.tz_localize(None)
        host_metric_dataframe['last_automation'] = pd.to_datetime(
            host_metric_dataframe['last_automation']).dt.tz_localize(None)
        host_metric_dataframe['last_deleted'] = pd.to_datetime(
            host_metric_dataframe['last_deleted']).dt.tz_localize(None)

        host_metric_dataframe['days_automated'] = (
            host_metric_dataframe['last_automation'] - host_metric_dataframe['first_automation']).dt.days
        host_metric_dataframe['days_automated'][host_metric_dataframe['days_automated'] < 0] = 0

        if self.extra_params.get("opt_ephemeral") is not None:
            ephemeral_usage_dataframe = self.compute_ephemeral_intervals(self.df_managed_nodes_query(host_metric_dataframe, ephemeral=True))

        # Create the workbook and worksheets
        self.wb.remove(self.wb.active) # delete the default sheet
        self.wb.create_sheet(title="Usage Reporting")

        # First sheet with billing
        ws = self.wb.worksheets[0]

        self._init_dimensions(ws)
        current_row = 1
        current_row = self._build_heading_h1(1, ws)
        current_row = self._build_header(current_row, ws)

        current_row = self._build_updated_timestamp(current_row, ws)
        current_row = self._build_data_section(current_row, ws, host_metric_dataframe, ephemeral_usage_dataframe)

        # Add optional sheets
        sheet_index = 1
        if "managed_nodes" in self.optional_report_sheets():
            # Sheet with list of managed nodes
            if self.extra_params.get("opt_ephemeral") is None:
                self.wb.create_sheet(title="Managed nodes")
                ws = self.wb.worksheets[sheet_index]
                current_row = self._build_data_section_host_metrics(
                    1, ws, self.df_managed_nodes_query(host_metric_dataframe))
                sheet_index += 1
            else:
                self.wb.create_sheet(title="Managed nodes")
                ws = self.wb.worksheets[sheet_index]
                current_row = self._build_data_section_host_metrics(
                    1, ws, self.df_managed_nodes_query(host_metric_dataframe, ephemeral=False))
                sheet_index += 1

                self.wb.create_sheet(title="Managed nodes ephemeral")
                ws = self.wb.worksheets[sheet_index]
                current_row = self._build_data_section_host_metrics(
                    1, ws, self.df_managed_nodes_query(host_metric_dataframe, ephemeral=True))
                sheet_index += 1

                self.wb.create_sheet(title="Managed nodes ephemeral usage")
                ws = self.wb.worksheets[sheet_index]
                current_row = self._build_data_section_ephemeral_usage(
                    1, ws, ephemeral_usage_dataframe)
                sheet_index += 1

            self.wb.create_sheet(title="Deleted Managed nodes")
            ws = self.wb.worksheets[sheet_index]
            current_row = self._build_data_section_host_metrics(
                1, ws, self.df_deleted_managed_nodes_query(host_metric_dataframe))
            sheet_index += 1

        return self.wb

    def df_managed_nodes_query(self, dataframe, ephemeral=None):
        if ephemeral is None:
            return dataframe[dataframe["deleted"]==False]
        else:
            # Take only non deleted
            dataframe = dataframe[dataframe["deleted"]==False]

            # Filter ephemeral based on number of automated days
            ephemeral_days = parse_number_of_days(self.extra_params.get("opt_ephemeral"))

            # Ephemeral threshold, host's first automation must be older than ephemeral threshold
            # to be considered as ephemeral
            ephemeral_threshold = pd.to_datetime(datetime.datetime.now() - datetime.timedelta(days=ephemeral_days - 1)).replace(
                hour=0, minute=0, second=0, microsecond=0).tz_localize(None)

            if ephemeral is True:
                return dataframe[(dataframe["days_automated"] <= ephemeral_days) &
                                 (dataframe["first_automation"] <= ephemeral_threshold)]
            if ephemeral is False:
                return dataframe[(dataframe["days_automated"] > ephemeral_days) |
                                 (dataframe["first_automation"] > ephemeral_threshold)]

    def df_deleted_managed_nodes_query(self, dataframe):
        return dataframe[dataframe["deleted"]==True]

    def get_intervals(self, start_date, end_date, interval_size):
        intervals = []
        current_start = start_date
        current_end = current_start + datetime.timedelta(days=interval_size) - datetime.timedelta(microseconds=1)

        while current_end <= end_date:
            intervals.append((current_start, current_end))

            current_start += datetime.timedelta(days=1)
            current_end = current_start + datetime.timedelta(days=interval_size) - datetime.timedelta(microseconds=1)

        return intervals

    def compute_ephemeral_intervals(self, host_metric_dataframe):
        # Convert input date strings to datetime objects
        start_date = pd.to_datetime(
            self.extra_params['since_date']).tz_localize(None)
        end_date = pd.to_datetime(
            self.extra_params['until_date']).tz_localize(None) + datetime.timedelta(days=1) - datetime.timedelta(microseconds=1)

        ephemeral_days = parse_number_of_days(self.extra_params.get("opt_ephemeral"))
        ephemeral_usage_intervals = []
        intervals = self.get_intervals(start_date, end_date, ephemeral_days)
        for window_start, window_end in intervals:
            print(f"Processing {window_start}, {window_end}")
            filtered = host_metric_dataframe[(host_metric_dataframe["last_automation"] >= window_start) & (host_metric_dataframe["first_automation"] <= window_end)]
            ephemeral_usage_intervals.append({
                "window_start": window_start,
                "window_end": window_end,
                "ephemeral_hosts": filtered["hostname"].nunique(),
            })

        return pd.DataFrame(ephemeral_usage_intervals)


    def _init_dimensions(self, ws):
        for key, value in self.config['column_widths'].items():
            ws.column_dimensions[get_column_letter(key)].width = value

    def _build_data_section(self, current_row, ws, dataframe, ephemeral_usage_dataframe):
        header_font = Font(name=self.FONT,
                           size=10,
                           color=self.BLACK_COLOR_HEX,
                           bold=True)
        value_font = Font(name=self.FONT,
                          size=10,
                          color=self.BLACK_COLOR_HEX)

        dotted_border = Border(
            left=Side(border_style='dotted', color=self.BLACK_COLOR_HEX),
            right=Side(border_style='dotted', color=self.BLACK_COLOR_HEX),
            top=Side(border_style='dotted', color=self.BLACK_COLOR_HEX),
            bottom=Side(border_style='dotted', color=self.BLACK_COLOR_HEX),
        )

        second_line_dotted_border = Border(
            left=Side(border_style='dotted', color=self.BLACK_COLOR_HEX),
            right=Side(border_style='dotted', color=self.BLACK_COLOR_HEX),
            bottom=Side(border_style='dotted', color=self.BLACK_COLOR_HEX),
        )

        header_border = Border(
            left=Side(border_style='medium', color=self.BLACK_COLOR_HEX),
            right=Side(border_style='medium', color=self.BLACK_COLOR_HEX),
            top=Side(border_style='medium', color=self.BLACK_COLOR_HEX),
            bottom=Side(border_style='medium', color=self.BLACK_COLOR_HEX),
        )

        ccsp_report = []

        if self.extra_params.get("opt_ephemeral") is None:
            # Automated hosts
            ccsp_report_item = {
                'description': "Automated hosts",
                'quantity_consumed': self.df_managed_nodes_query(dataframe)["hostname"].nunique()
            }
            ccsp_report.append(ccsp_report_item)
        else:
            # Automated hosts non ephemeral
            ccsp_report_item = {
                'description': "Automated hosts",
                'quantity_consumed': self.df_managed_nodes_query(dataframe, ephemeral=False)["hostname"].nunique()
            }
            ccsp_report.append(ccsp_report_item)

            # Automated hosts ephemeral total
            ccsp_report_item = {
                'description': "Ephemeral automated hosts total",
                'quantity_consumed': self.df_managed_nodes_query(dataframe, ephemeral=True)["hostname"].nunique()
            }
            ccsp_report.append(ccsp_report_item)

            # Ephemeral automated hosts maximum concurrent usage in defined interval"
            ccsp_report_item = {
                'description': "Ephemeral automated hosts maximum\nconcurrent usage in defined interval",
                'quantity_consumed': ephemeral_usage_dataframe['ephemeral_hosts'].max()
            }
            ccsp_report.append(ccsp_report_item)

        # Deleted automated hosts
        ccsp_report_item = {
            'description': "Deleted Automated hosts",
            'quantity_consumed': self.df_deleted_managed_nodes_query(dataframe)["hostname"].nunique()
        }
        ccsp_report.append(ccsp_report_item)

        ccsp_report = pd.DataFrame(ccsp_report)

        # order the columns right
        ccsp_report = ccsp_report.reset_index()
        ccsp_report = ccsp_report.reindex(columns=['description',
                                                   'quantity_consumed'])

        # Rename the columns based on the template
        ccsp_report_dataframe = ccsp_report.rename(
            columns={"description": "Description",
                     "quantity_consumed": "Quantity",
                    })

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            if row_counter == 0:
                rd = ws.row_dimensions[r_idx]
                rd.height = 35
            elif row_counter >= 1:
                # Set bigger height of the data columns
                rd = ws.row_dimensions[r_idx]
                rd.height = 25

            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    cell.border = header_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    if row_counter == 1:
                        # set value style
                        cell.font = value_font
                        cell.border = second_line_dotted_border
                    else:
                        # set value style
                        cell.font = value_font
                        cell.border = dotted_border

            row_counter += 1

        return current_row + row_counter

    def _build_data_section_host_metrics(self, current_row, ws, dataframe):
        for key, value in self.config['data_column_widths'].items():
            ws.column_dimensions[get_column_letter(key)].width = value

        header_font = Font(name=self.FONT,
                           size=10,
                           color=self.BLACK_COLOR_HEX,
                           bold=True)
        value_font = Font(name=self.FONT,
                          size=10,
                          color=self.BLACK_COLOR_HEX)

        # Rename the columns based on the template
        ccsp_report_dataframe = dataframe.reset_index()
        ccsp_report_dataframe = ccsp_report_dataframe.reindex(
            columns=[
                'hostname',
                'first_automation',
                'last_automation',
                'automated_counter',
                'days_automated',
                'deleted_counter',
                'last_deleted',
            ]
        )

        ccsp_report_dataframe = ccsp_report_dataframe.rename(
            columns={
                'hostname': "Host name",
                'first_automation': "First\nautomation",
                'last_automation': "Last\nautomation",
                'automated_counter': "Number of\nAutomations",
                'days_automated': "Number of days\nbetween first_automation\nand last_automation",
                'deleted_counter': "Number of\nDeletions",
                'last_deleted': "Last\ndeleted",
            }
        )

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    rd = ws.row_dimensions[r_idx]
                    rd.height = 25
                else:
                    # set value style
                    cell.font = value_font

            row_counter += 1

        return current_row + row_counter

    def _build_data_section_ephemeral_usage (self, current_row, ws, dataframe):
        for key, value in self.config['uniform_column_widths'].items():
            ws.column_dimensions[get_column_letter(key)].width = value

        header_font = Font(name=self.FONT,
                           size=10,
                           color=self.BLACK_COLOR_HEX,
                           bold=True)
        value_font = Font(name=self.FONT,
                          size=10,
                          color=self.BLACK_COLOR_HEX)

        # Rename the columns based on the template
        ccsp_report_dataframe = dataframe.reset_index()
        ccsp_report_dataframe = ccsp_report_dataframe.reindex(
            columns=[
                'window_start',
                'window_end',
                'ephemeral_hosts',
            ]
        )

        ccsp_report_dataframe = ccsp_report_dataframe.rename(
            columns={
                'window_start': "Start of the\nephemeral window",
                'window_end': "Start of the\nephemeral window",
                'ephemeral_hosts': "Ephemeral automated hosts",
            }
        )

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    rd = ws.row_dimensions[r_idx]
                    rd.height = 25
                else:
                    # set value style
                    cell.font = value_font

            row_counter += 1

        return current_row + row_counter

    def _build_heading_h1(self, current_row, ws):
        # Merge cells and insert the h1 heading
        ws.merge_cells(start_row=current_row, start_column=1,
                            end_row=current_row, end_column=2)
        h1_heading_cell = ws.cell(row=current_row, column=1)
        h1_heading_cell.value = self.config['h1_heading']['value']

        h1_heading_cell.font = Font(name=self.FONT,
                                    size=12,
                                    bold=True,
                                    )

        current_row += 1
        return current_row

    def _build_updated_timestamp(self, current_row, ws):
        cell = ws.cell(row=1, column=4)
        cell.value = f"Updated: {time.strftime('%b %d, %Y')}"

        return current_row

    def _build_header(self, current_row, ws):
        # Insert the header
        for header_row in self.config['header']:
            header_label_font = Font(name=self.FONT,
                                     size=12,
                                     color=self.BLACK_COLOR_HEX)
            header_value_font = Font(name=self.FONT,
                                     size=12,
                                     color=self.BLACK_COLOR_HEX)

            cell = ws.cell(row=current_row, column=1)
            cell.value = header_row['label']
            cell.font = header_label_font

            cell = ws.cell(row=current_row, column=2)
            if header_row['label'] == "Report Period (YYYY-MM-DD, YYYY-MM-DD)":
                # Insert dynamic report period into the specific header column
                cell.fill = PatternFill("solid", fgColor=self.GREEN_COLOR_HEX)
                cell.value = self.extra_params['report_period_range']
            else:
                cell.fill = PatternFill("solid", fgColor=self.GREEN_COLOR_HEX)
                cell.value = header_row['value']
            cell.font = header_value_font
            current_row += 1

        return current_row