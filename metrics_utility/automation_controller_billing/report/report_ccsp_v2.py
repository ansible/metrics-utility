######################################
# Code for building the spreadsheet
######################################
from metrics_utility.automation_controller_billing.report.base import Base
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

import pandas as pd
import time

class ReportCCSPv2(Base):
    # BLACK_COLOR_HEX = "00000000"
    # WHITE_COLOR_HEX = "00FFFFFF"
    # BLUE_COLOR_HEX = "000000FF"
    # RED_COLOR_HEX = "FF0000"
    # LIGHT_BLUE_COLOR_HEX = "d4eaf3"
    # GREEN_COLOR_HEX = "92d050"
    # FONT = "Arial"
    # PRICE_FORMAT = '$#,##0.00'

    def __init__(self, dataframe, report_period, extra_params):
        self.wb = Workbook()

        self.dataframe = dataframe
        self.report_period = report_period
        self.extra_params = extra_params
        self.price_per_node = extra_params['price_per_node']

        self.config = {
            'sku': extra_params['report_sku'],
            'h1_heading': {
                'value': extra_params['report_h1_heading'],
            },
            'po_number': {
                'label': 'PO Number',
                'value': extra_params['report_po_number'],
            },
            'header': [{
                'label': 'CCSP Company Name',
                'value': extra_params['report_company_name'],
            },{
                'label': 'CCSP Email',
                'value': extra_params['report_email'],
            },{
                'label': 'CCSP RHN Login',
                'value': extra_params['report_rhn_login'],
            },{
                'label': 'Report Period (YYYY-MM)',
                'value': '<autogenerated>',
            }],
        }

        default_sku_description = [
            ['SKU',
            'SKU Description',
            '',
            'Term',
            'Unit of Measure',
            'Currency',
            'MSRP'],
            [f"{extra_params['report_sku']}",
            f"{extra_params['report_sku_description']}",
            '',
            'MONTH',
            'MANAGED NODE',
            'USD',
            f"{extra_params['price_per_node']}"]
        ]

        default_column_widths = {
            1: 40,
            2: 25,
            3: 15,
            4: 15,
            5: 15,
            6: 20,
            7: 15,
            8: 30,
            9: 20,
            10: 20,
            11: 20
        }

        default_data_column_widths = {
            1: 40,
            2: 20,
            3: 20,
            4: 20,
            5: 20,
            6: 20,
            7: 20
        }

        self.config['sku_description'] = default_sku_description
        self.config['column_widths'] = default_column_widths
        self.config['data_column_widths'] = default_data_column_widths

    def build_spreadsheet(self):
        # Fix host names in the event data, to take in account the variables
        job_host_sumary_dataframe = self.dataframe[0]
        events_dataframe = self.dataframe[1]
        events_dataframe = self._fix_event_host_names(job_host_sumary_dataframe, events_dataframe)

        # Create the workbook and worksheets
        self.wb.remove(self.wb.active) # delete the default sheet
        self.wb.create_sheet(title="Usage Reporting")

        # First sheet with billing
        ws = self.wb.worksheets[0]

        self._init_dimensions(ws)
        current_row = self._build_heading_h1(1, ws)
        current_row = self._build_header(current_row, ws)
        current_row = self._build_po_number(current_row, ws)
        current_row = self._build_updated_timestamp(current_row, ws)
        current_row = self._build_data_section(current_row, ws, job_host_sumary_dataframe)

        # Add optional sheets
        sheet_index = 1
        if "managed_nodes" in self.optional_report_sheets():
            # Sheet with list of managed nodes
            self.wb.create_sheet(title="Managed nodes")
            ws = self.wb.worksheets[sheet_index]
            current_row = self._build_data_section_usage_by_node(1, ws, job_host_sumary_dataframe)
            sheet_index += 1

        if "usage_by_organizations" in self.optional_report_sheets():
            # Sheet with usage by org
            self.wb.create_sheet(title="Usage by organizations")
            ws = self.wb.worksheets[sheet_index]
            current_row = self._build_data_section_usage_by_org(1, ws, job_host_sumary_dataframe)
            sheet_index += 1

        if events_dataframe is not None:
            if "usage_by_collections" in self.optional_report_sheets():
                # Sheet with usage by collections
                self.wb.create_sheet(title="Usage by collections")
                ws = self.wb.worksheets[sheet_index]
                current_row = self._build_data_section_usage_by_collections(1, ws, events_dataframe)
                sheet_index += 1

            if "usage_by_roles" in self.optional_report_sheets():
                # Sheet with usage by roles
                self.wb.create_sheet(title="Usage by roles")
                ws = self.wb.worksheets[sheet_index]
                current_row = self._build_data_section_usage_by_roles(1, ws, events_dataframe)
                sheet_index += 1

            if "usage_by_modules" in self.optional_report_sheets():
                # Sheet with usage by modules
                self.wb.create_sheet(title="Usage by modules")
                ws = self.wb.worksheets[sheet_index]
                current_row = self._build_data_section_usage_by_modules(1, ws, events_dataframe)
                sheet_index += 1

        if "managed_nodes_by_organizations" in self.optional_report_sheets():
            # Sheet with list of managed nodes by organization, this will generate multiple tabs
            organization_names = sorted(job_host_sumary_dataframe['organization_name'].unique())
            for organization_name in organization_names:
                self.wb.create_sheet(title=organization_name)
                ws = self.wb.worksheets[sheet_index]

                # Filter the data for a certain organization
                filtered_job_host_sumary_dataframe = job_host_sumary_dataframe[
                    job_host_sumary_dataframe["organization_name"] == organization_name]
                current_row = self._build_data_section_usage_by_node(1, ws, filtered_job_host_sumary_dataframe, mode="by_organization")
                sheet_index += 1

        return self.wb

    def _build_data_section_usage_by_org(self, current_row, ws, dataframe):
        for key, value in self.config['data_column_widths'].items():
            ws.column_dimensions[get_column_letter(key)].width = value

        header_font = Font(name=self.FONT,
                           size=10,
                           color=self.BLACK_COLOR_HEX,
                           bold=True)
        value_font = Font(name=self.FONT,
                          size=10,
                          color=self.BLACK_COLOR_HEX)

        # Rename the columns based on the template
        ccsp_report_dataframe = (
            dataframe.groupby('organization_name', dropna=False)
            .agg(
                host_runs_unique=('host_name', 'nunique'),
                host_runs=('host_name', 'count'),
                task_runs=('task_runs', 'sum')
            )
        )
        ccsp_report_dataframe = ccsp_report_dataframe.reset_index()
        ccsp_report_dataframe = ccsp_report_dataframe.reindex(
            columns=[
                'organization_name',
                'host_runs_unique',
                'host_runs',
                'task_runs'
            ]
        )

        ccsp_report_dataframe = ccsp_report_dataframe.rename(
            columns={
                "organization_name": "Organization name",
                "host_runs_unique": "Unique managed nodes\nautomated",
                "host_runs": "Non-unique managed\nnodes automated",
                "task_runs": "Number of task\nruns",
            }
        )

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    rd = ws.row_dimensions[r_idx]
                    rd.height = 25
                else:
                    # set value style
                    cell.font = value_font

            row_counter += 1

        return current_row + row_counter

    def _init_dimensions(self, ws):
        for key, value in self.config['column_widths'].items():
            ws.column_dimensions[get_column_letter(key)].width = value

    def _build_heading_h1(self, current_row, ws):
        # Merge cells and insert the h1 heading
        ws.merge_cells(start_row=current_row, start_column=1,
                            end_row=current_row, end_column=2)
        h1_heading_cell = ws.cell(row=current_row, column=1)
        h1_heading_cell.value = self.config['h1_heading']['value']

        # h1_heading_cell.fill = PatternFill("solid", fgColor=self.BLACK_COLOR_HEX)
        h1_heading_cell.font = Font(name=self.FONT,
                                    size=12,
                                    bold=True,
                                    )#color=self.WHITE_COLOR_HEX)

        # h1_heading_cell.alignment = Alignment(horizontal='center')

        current_row += 1
        return current_row

    def _build_updated_timestamp(self, current_row, ws):
        cell = ws.cell(row=1, column=8)
        cell.value = f"Updated: {time.strftime('%b %d, %Y')}"

        return current_row

    def _build_po_number(self, current_row, ws):
        # Add the h2 heading payment heading
        green_background = PatternFill("solid", fgColor=self.GREEN_COLOR_HEX)

        # PO number heading and value with green background
        cell = ws.cell(row=4, column=5)
        cell.value = self.config['po_number']['label']
        cell.fill = green_background

        cell = ws.cell(row=4, column=6)
        cell.value = self.config['po_number']['value']
        cell.fill = green_background

        return current_row

    def _build_header(self, current_row, ws):
        # Insert the header
        for header_row in self.config['header']:
            header_label_font = Font(name=self.FONT,
                                     size=12,
                                     color=self.BLACK_COLOR_HEX)
            header_value_font = Font(name=self.FONT,
                                     size=12,
                                     color=self.BLACK_COLOR_HEX)

            cell = ws.cell(row=current_row, column=1)
            cell.value = header_row['label']
            cell.font = header_label_font

            cell = ws.cell(row=current_row, column=2)
            if header_row['label'] == "Report Period (YYYY-MM)":
                # Insert dynamic report period into the specific header column
                cell.fill = PatternFill("solid", fgColor=self.GREEN_COLOR_HEX)
                cell.value = self.report_period
            else:
                cell.fill = PatternFill("solid", fgColor=self.GREEN_COLOR_HEX)
                cell.value = header_row['value']
            cell.font = header_value_font
            current_row += 1

        return current_row

    def _build_sku_description(self, current_row, ws):
        # Insert the header
        row_counter = 0
        for header_row in self.config['sku_description']:
            header_font = Font(name=self.FONT,
                               size=11,
                               color=self.BLACK_COLOR_HEX,
                               bold=True)
            header_border = Border(left=Side(border_style='thin',
                                             color=self.BLACK_COLOR_HEX),
                                   right=Side(border_style='thin',
                                              color=self.BLACK_COLOR_HEX),
                                   top=Side(border_style='thin',
                                            color=self.BLACK_COLOR_HEX),
                                   bottom=Side(border_style='thin',
                                               color=self.BLACK_COLOR_HEX))
            value_font = Font(name=self.FONT,
                              size=11,
                              color=self.BLACK_COLOR_HEX)
            col_counter = 0
            for col_value in header_row:
                col_counter += 1

                cell = ws.cell(row=current_row + row_counter, column=col_counter)
                cell.value = col_value

                if row_counter == 0:
                    # header
                    cell.font = header_font
                    cell.border = header_border
                else:
                    # row
                    cell.font = value_font

            row_counter +=1
        current_row = current_row + row_counter
        # make extra 1 row space
        current_row += 1

        return current_row

    def _build_data_section(self, current_row, ws, dataframe):
        header_font = Font(name=self.FONT,
                           size=10,
                           color=self.BLACK_COLOR_HEX,
                           bold=True)
        value_font = Font(name=self.FONT,
                          size=10,
                          color=self.BLACK_COLOR_HEX)

        dotted_border = Border(
            left=Side(border_style='dotted', color=self.BLACK_COLOR_HEX),
            right=Side(border_style='dotted', color=self.BLACK_COLOR_HEX),
            top=Side(border_style='dotted', color=self.BLACK_COLOR_HEX),
            bottom=Side(border_style='dotted', color=self.BLACK_COLOR_HEX),
        )

        second_line_dotted_border = Border(
            left=Side(border_style='dotted', color=self.BLACK_COLOR_HEX),
            right=Side(border_style='dotted', color=self.BLACK_COLOR_HEX),
            bottom=Side(border_style='dotted', color=self.BLACK_COLOR_HEX),
        )

        header_border = Border(
            left=Side(border_style='medium', color=self.BLACK_COLOR_HEX),
            right=Side(border_style='medium', color=self.BLACK_COLOR_HEX),
            top=Side(border_style='medium', color=self.BLACK_COLOR_HEX),
            bottom=Side(border_style='medium', color=self.BLACK_COLOR_HEX),
        )

        ccsp_report = {}
        quantity_consumed = dataframe["host_name"].nunique()
        if quantity_consumed > 0:
            # COmpute the unique hostnam count that are in the df index
            ccsp_report["end_user_company_name"] = self.extra_params['report_end_user_company_name']
            ccsp_report["end_user_company_city"] = self.extra_params['report_end_user_company_city']
            ccsp_report["end_user_company_state"] = self.extra_params['report_end_user_company_state']
            ccsp_report["end_user_company_country"] = self.extra_params['report_end_user_company_country']

            ccsp_report["quantity_consumed"] = quantity_consumed
            ccsp_report['mark_x'] = ''
            ccsp_report['sku_number'] = self.extra_params['report_sku']
            ccsp_report['sku_description'] = self.extra_params['report_sku_description']
            ccsp_report['notes'] = ''

            ccsp_report['unit_price'] = round(self.price_per_node, 2)
            ccsp_report['extended_unit_price'] = round((ccsp_report['quantity_consumed'] * ccsp_report['unit_price']), 2)
            ccsp_report = pd.DataFrame([ccsp_report])

            # order the columns right
            ccsp_report = ccsp_report.reset_index()
            ccsp_report = ccsp_report.reindex(columns=['end_user_company_name',
                                                       'mark_x',
                                                       'end_user_company_city',
                                                       'end_user_company_state',
                                                       'end_user_company_country',
                                                       'sku_number',
                                                       'quantity_consumed',
                                                       'sku_description',
                                                       'unit_price',
                                                       'extended_unit_price',
                                                       'notes'])

        else:
            # Generate empty df if there were no billing data
            ccsp_report = pd.DataFrame([ccsp_report])

        # Rename the columns based on the template
        ccsp_report_dataframe = ccsp_report.rename(
            columns={"end_user_company_name": "End User Company Name",
                     "mark_x": "Enter 'X' to indicate\nInteral Usage",
                     "end_user_company_city": "End User\nCity",
                     "end_user_company_state": "End User\nState/Prov",
                     "end_user_company_country": "Country Where\nSKU Consumed",
                     "sku_number": "SKU Number",
                     "quantity_consumed": "Quantity",
                     "sku_description": "SKU Description",
                     "unit_price": "SKU Unit Price",
                     "extended_unit_price": "SKU Extended Unit\nPrice",
                     "notes": "Notes",
                    })

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            if row_counter == 0:
                rd = ws.row_dimensions[r_idx]
                rd.height = 35
            elif row_counter >= 1:
                # Set bigger height of the data columns
                rd = ws.row_dimensions[r_idx]
                rd.height = 25

            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    cell.border = header_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    if row_counter == 1:
                        # set value style
                        cell.font = value_font
                        cell.border = second_line_dotted_border
                    else:
                        # set value style
                        cell.font = value_font
                        cell.border = dotted_border

                    if c_idx >= 8 and c_idx <= 11:
                        cell.fill = PatternFill("solid", fgColor=self.LIGHT_BLUE_COLOR_HEX)
                    if c_idx >= 9 and c_idx <= 10:
                        # Format all price cols
                        cell.number_format = self.PRICE_FORMAT
                    if c_idx == 10:
                        # Override the value of the extended price (number of nodes X unitp rice)
                        # Multiply columns 3x4 instead of inserting the price per org
                        cell_m_1 = ws.cell(row=r_idx, column=7).column_letter + str(r_idx)
                        cell_m_2 = ws.cell(row=r_idx, column=9).column_letter + str(r_idx)
                        cell.value = '={0}*{1}'.format(cell_m_1, cell_m_2)

            row_counter += 1

        # Generate 5 more blank rows at the end
        for r_counter in range(5):
            # Set bigger height of the data columns
            r_idx = current_row + row_counter
            rd = ws.row_dimensions[r_idx]
            rd.height = 25

            for c_counter in range(11):
                c_idx = c_counter + 1
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = dotted_border
                cell.font = value_font

                if c_idx >= 8 and c_idx <= 11:
                    cell.fill = PatternFill("solid", fgColor=self.LIGHT_BLUE_COLOR_HEX)
                if c_idx >= 9 and c_idx <= 10:
                    # Format all price cols
                    cell.number_format = self.PRICE_FORMAT
                if c_idx == 10:
                    # Override the value of the extended price (number of nodes X unitp rice)
                    # Multiply columns 3x4 instead of inserting the price per org
                    cell_m_1 = ws.cell(row=r_idx, column=7).column_letter + str(r_idx)
                    cell_m_2 = ws.cell(row=r_idx, column=9).column_letter + str(r_idx)
                    cell.value = '={0}*{1}'.format(cell_m_1, cell_m_2)

            row_counter += 1

        if row_counter >= 2:
            # If there is at least 1 data column, insert the sum:
            first_row = current_row + 1 # ignore the header
            last_row = current_row + row_counter - 1

            # Sum description
            cell = ws.cell(row=2, column=9)
            cell.value = "Grand total"
            cell.fill = PatternFill("solid", fgColor=self.LIGHT_BLUE_COLOR_HEX)

            # Sum value
            cell = ws.cell(row=2, column=10)
            cell_sum_start = cell.column_letter + str(first_row)
            cell_sum_end = cell.column_letter + str(last_row)
            cell.value = '=SUM({0}:{1})'.format(cell_sum_start, cell_sum_end)
            cell.fill = PatternFill("solid", fgColor=self.LIGHT_BLUE_COLOR_HEX)
            cell.font = Font(name=self.FONT,
                             size=10,
                             color=self.RED_COLOR_HEX)

        return current_row + row_counter
