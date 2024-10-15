######################################
# Code for building the spreadsheet
######################################
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

import os
import pandas as pd


class Base:
    BLACK_COLOR_HEX = "00000000"
    WHITE_COLOR_HEX = "00FFFFFF"
    BLUE_COLOR_HEX = "000000FF"
    RED_COLOR_HEX = "FF0000"
    LIGHT_BLUE_COLOR_HEX = "d4eaf3"
    GREEN_COLOR_HEX = "92d050"
    FONT = "Arial"
    PRICE_FORMAT = '$#,##0.00'

    @staticmethod
    def optional_report_sheets():
        return os.environ.get(
            'METRICS_UTILITY_OPTIONAL_CCSP_REPORT_SHEETS',
            'managed_nodes,usage_by_organizations,usage_by_collections,usage_by_roles,'\
            'usage_by_modules').split(",")

    def _fix_event_host_names(self, mapping_dataframe, destination_dataframe):
        if destination_dataframe is None:
            return None

        def concatenate_columns_mapping(row):
            return f"{row['original_host_name']}__{str(row['install_uuid'])}__{str(row['job_remote_id'])}"

        def concatenate_columns_destination(row):
            return f"{row['host_name']}__{str(row['install_uuid'])}__{str(row['job_remote_id'])}"

        # Apply the function to each row of the DataFrame
        mapping_dataframe['host_composite_id'] = mapping_dataframe.apply(concatenate_columns_mapping, axis=1)
        mapping_dataframe = mapping_dataframe.set_index("host_composite_id")
        mapping_dataframe = mapping_dataframe["host_name"].astype(str).to_dict()

        def apply_mapping(row):
            return mapping_dataframe.get(f"{row['host_name']}__{str(row['install_uuid'])}__{row['job_remote_id']}", row['host_name'])

        destination_dataframe['host_name'] = destination_dataframe.apply(apply_mapping, axis=1)
        destination_dataframe['host_composite_id'] = destination_dataframe.apply(concatenate_columns_destination, axis=1)

        return destination_dataframe

    def _build_data_section_usage_by_node(self, current_row, ws, dataframe, mode=None):
        for key, value in self.config['data_column_widths'].items():
            ws.column_dimensions[get_column_letter(key)].width = value

        header_font = Font(name=self.FONT,
                           size=10,
                           color=self.BLACK_COLOR_HEX,
                           bold=True)
        value_font = Font(name=self.FONT,
                          size=10,
                          color=self.BLACK_COLOR_HEX)

        # Rename the columns based on the template
        ccsp_report_dataframe = (
            dataframe.groupby('host_name', dropna=False)
            .agg(
                organizations=('organization_name', 'nunique'),
                host_runs=('host_name', 'count'),
                task_runs=('task_runs', 'sum'),
                first_automation=('first_automation', 'min'),
                last_automation=('last_automation', 'max')
            )
        )
        ccsp_report_dataframe = ccsp_report_dataframe.reset_index()
        columns = [
            'host_name',
            'organizations',
            'host_runs',
            'task_runs',
            'first_automation',
            'last_automation',
        ]
        if mode == "by_organization":
            # Filter some columns out based on mode
            columns = [col for col in columns if col not in ['organizations']]
        ccsp_report_dataframe = ccsp_report_dataframe.reindex(
            columns=columns
        )

        labels = {
            "host_name": "Host name",
            "organizations": "Automated by\norganizations",
            "host_runs": "Non-unique managed\nnodes automated",
            "task_runs": "Number of task\nruns",
            'first_automation': "First\nautomation",
            'last_automation': "Last\nautomation",
        }
        labels = {k:v for k, v in labels.items() if k in columns}
        ccsp_report_dataframe = ccsp_report_dataframe.rename(
            columns=labels
        )

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    rd = ws.row_dimensions[r_idx]
                    rd.height = 25
                else:
                    # set value style
                    cell.font = value_font

            row_counter += 1

        return current_row + row_counter

    def _build_data_section_usage_by_collections(self, current_row, ws, dataframe):
        for key, value in self.config['data_column_widths'].items():
            ws.column_dimensions[get_column_letter(key)].width = value

        header_font = Font(name=self.FONT,
                           size=10,
                           color=self.BLACK_COLOR_HEX,
                           bold=True)
        value_font = Font(name=self.FONT,
                          size=10,
                          color=self.BLACK_COLOR_HEX)

        # Take the content explorer dataframe and extract specific group by
        ccsp_report_dataframe = dataframe.groupby(
            ["collection_name"], dropna=False
        ).agg(
            host_runs_unique=('host_name', 'nunique'),
            host_runs=('host_composite_id', 'nunique'),
            task_runs=('task_runs', 'sum'),
            duration=('duration', "sum"))

        # Rename the columns based on the template
        ccsp_report_dataframe = ccsp_report_dataframe.reset_index()

        ccsp_report_dataframe = ccsp_report_dataframe.rename(
            columns={
                "collection_name": "Collection name",
                "host_runs_unique": "Unique managed nodes\nautomated",
                "host_runs": "Non-unique managed\nnodes automated",
                "task_runs": "Number of task\nruns",
                "duration": "Duration of task\nruns [seconds]",
            }
        )

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    rd = ws.row_dimensions[r_idx]
                    rd.height = 25
                else:
                    # set value style
                    cell.font = value_font

            row_counter += 1

        return current_row + row_counter

    def _build_data_section_usage_by_roles(self, current_row, ws, dataframe):
        for key, value in self.config['data_column_widths'].items():
            ws.column_dimensions[get_column_letter(key)].width = value

        header_font = Font(name=self.FONT,
                           size=10,
                           color=self.BLACK_COLOR_HEX,
                           bold=True)
        value_font = Font(name=self.FONT,
                          size=10,
                          color=self.BLACK_COLOR_HEX)

        # Take the content explorer dataframe and extract specific group by
        ccsp_report_dataframe = dataframe.groupby(
            ["role_name"], dropna=False
        ).agg(
            host_runs_unique=('host_name', 'nunique'),
            host_runs=('host_composite_id', 'nunique'),
            task_runs=('task_runs', 'sum'),
            duration=('duration', "sum"))

        # Rename the columns based on the template
        ccsp_report_dataframe = ccsp_report_dataframe.reset_index()

        ccsp_report_dataframe = ccsp_report_dataframe.rename(
            columns={
                "role_name": "Role name",
                "host_runs_unique": "Unique managed nodes\nautomated",
                "host_runs": "Non-unique managed\nnodes automated",
                "task_runs": "Number of task\nruns",
                "duration": "Duration of task\nruns [seconds]",
            }
        )

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                # cell.border = dotted_border

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    rd = ws.row_dimensions[r_idx]
                    rd.height = 25
                else:
                    # set value style
                    cell.font = value_font

            row_counter += 1

        return current_row + row_counter

    def _build_data_section_usage_by_modules(self, current_row, ws, dataframe):
        for key, value in self.config['data_column_widths'].items():
            ws.column_dimensions[get_column_letter(key)].width = value

        header_font = Font(name=self.FONT,
                           size=10,
                           color=self.BLACK_COLOR_HEX,
                           bold=True)
        value_font = Font(name=self.FONT,
                          size=10,
                          color=self.BLACK_COLOR_HEX)

        ccsp_report_dataframe = dataframe.groupby(
            ["module_name"], dropna=False
        ).agg(
            host_runs_unique=('host_name', 'nunique'),
            host_runs=('host_composite_id', 'nunique'),
            task_runs=('task_runs', 'sum'),
            duration=('duration', "sum"))

        # Rename the columns based on the template
        ccsp_report_dataframe = ccsp_report_dataframe.reset_index()

        ccsp_report_dataframe = ccsp_report_dataframe.rename(
            columns={
                "module_name": "Module name",
                "host_runs_unique": "Unique managed nodes\nautomated",
                "host_runs": "Non-unique managed\nnodes automated",
                "task_runs": "Number of task\nruns",
                "duration": "Duration of task\nruns [seconds]",
            }
        )

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    rd = ws.row_dimensions[r_idx]
                    rd.height = 25
                else:
                    # set value style
                    cell.font = value_font

            row_counter += 1

        return current_row + row_counter
